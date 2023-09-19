__author__    = "Dr Pham Quang Phuong","Cao Anh Quoc Hung"
__copyright__ = "Copyright 2023"
__license__   = "All rights reserved"
__email__     = "phuong.phamquang@hust.edu.vn","caoanhquochung@gmail.com"
__status__    = "Released"
__version__   = "1.2"
"""
about: ....
"""
from new_kernel import add2CSV,PowerFlow,Parameter,Configuration
import os,sys,time
import openpyxl
import argparse
import random
import math
from pyswarms.utils.plotters import plot_cost_history
import pygad
import matplotlib.pyplot as plt


PARSER_INPUTS = argparse.ArgumentParser(epilog= "")
PARSER_INPUTS.usage = 'Distribution network analysis Tools'
PARSER_INPUTS.add_argument('-fi' , help = '*(str) Input file path .xlsx' , default = '',type=str,metavar='')
PARSER_INPUTS.add_argument('-fo' , help = ' (str) Output file path .csv' , default = '',type=str,metavar='')
ARGVS = PARSER_INPUTS.parse_known_args()[0]
#
#
def monteCarlo(nIter,lineOff0=[],shuntOff0=[]):
    nIter = int(nIter)
    print('Running monteCarlo nIter=%i'%nIter)
    param = Parameter(ARGVS.fi)
    rs = [[],[time.ctime(),'MonteCarlo init_pos lineOff0=%s shuntOff0=%s '%(str(lineOff0),str(shuntOff0),)],['iter','Objective','DeltaA','RateMax[%]','Umax[pu]','Umin[pu]','LineOff','ShuntOff'] ]
    add2CSV(ARGVS.fo,rs,',')
    r0 = {'Objective':math.inf,'LineOff':lineOff0,'ShuntOff':shuntOff0}
    if lineOff0 or shuntOff0:
        config = Configuration(param=param,lineOff=lineOff0,shuntOff=shuntOff0)
        pf = PowerFlow(config)
        r1 = pf.run1Config_WithObjective()
        rs = ['init','%.5f'%r1['Objective'],'%.5f'%r1['DeltaA'].real,'%.3f'%r1['RateMax[%]'],'%.3f'%r1['Umax[pu]'],'%.3f'%r1['Umin[pu]'],str(r1['LineOff']),str(r1['ShuntOff'])]
        add2CSV(ARGVS.fo,[rs],',')
        r0 = r1
    #
    for i in range(nIter):
        x = [random.randint(0,1) for _ in range(param.nVar)]
        config = Configuration(param=param,varFlag=x)
        pf = PowerFlow(config)
        r1 = pf.run1Config_WithObjective()
        if r1['FLAG']=='CONVERGENCE':
            if r1['Objective']<r0['Objective']:
                rs = [str(i),'%.5f'%r1['Objective'],'%.5f'%r1['DeltaA'].real,'%.3f'%r1['RateMax[%]'],'%.3f'%r1['Umax[pu]'],'%.3f'%r1['Umin[pu]'],str(r1['LineOff']),str(r1['ShuntOff'])]
                add2CSV(ARGVS.fo,[rs],',')
                r0 = r1
    #
    s1 = ['MonteCarlo','nIter',nIter,' time[s]',str(r0['LineOff']),str(r0['ShuntOff'])]
    add2CSV(ARGVS.fo,[s1],',')
    print('\nOutFile: '+os.path.abspath(ARGVS.fo))
    print(s1)
##    print(pf.nn)
#
class PSO:
    def func1(self,x,param):
        if type(x).__name__=='ndarray':
            x = x.tolist()[0]
        config = Configuration(param=param,varFlag=x)
        pf = PowerFlow(config)
        return pf.run1Config_WithObjective()['Objective']
    
    def run(self,nIter,lineOff0=[],shuntOff0=[]):
        nIter = int(nIter)
        print('Running PSO nIter=%i'%nIter)
        import pyswarms as ps
        import numpy as np
        #
        param = Parameter(ARGVS.fi)
        #
        config = Configuration(param,lineOff0,shuntOff0)

        options = {'c1': 0.5, 'c2': 0.3, 'w': 0.9,'k':1,'p':1}
        #
        pos0 = None
        if lineOff0 or shuntOff0:
            pos0 = np.array([config.getVarFlag()])
        #
        op = ps.discrete.binary.BinaryPSO(n_particles=1,dimensions=param.nVar,options=options,init_pos=pos0)
        cost_history = op.cost_history
        #
        cost, pos = op.optimize(self.func1, iters=nIter,param=param)
        #
        plot_cost_history(cost_history)
        plt.show()
        #
        config = Configuration(param=param,varFlag=pos)
        pf = PowerFlow(config)
        r1 = pf.run1Config_WithObjective()
        rs = [[],[time.ctime(),'PSO init_pos (lineOff)',str(lineOff0)],['Objective','DeltaA','RateMax[%]','Umax[pu]','Umin[pu]','LineOff','ShuntOff'] ]
        rs.append( ['%.5f'%r1['Objective'],'%.5f'%r1['DeltaA'].real,'%.3f'%r1['RateMax[%]'],'%.3f'%r1['Umax[pu]'],'%.3f'%r1['Umin[pu]'],str(r1['LineOff']),str(r1['ShuntOff'])])
        add2CSV(ARGVS.fo,rs,',')
        #
        s1 = ['PSO','nIter',nIter,str(r1['LineOff']),str(r1['ShuntOff'])]
        add2CSV(ARGVS.fo,[s1],',')
        print('\nOutFile: '+os.path.abspath(ARGVS.fo))
        print(s1)

class GA():
    def __init__(self,fi,fo):
        self.fo = fo
        self.fi = fi
        self.param = Parameter(fi)
        wbInput = openpyxl.load_workbook(os.path.abspath(fi),data_only=True)
        self.gaparam = self.__readParameter__(wbInput)
    def __readParameter__(self,wbInput):
        ws = wbInput['SETTING']
        k = 0
        res = {}
        while True:
            k+=1
            s1= ws.cell(k,1).value
            if type(s1)==str and s1.replace(' ','')=='##GAParameters':
                for j in range(1,100):
                    s2 = str(ws.cell(k+1,j).value).strip()
                    if s2=='None':
                        break
                    sa = str(ws.cell(k+2,j).value).split(',')
                    if len(sa)==1:
                        try:
                            res[s2] = float(sa[0])
                        except:
                            res[s2] = sa[0]
                    else:
                        res[s2] = [float(si) for si in sa]
                break
        return res
    def fitness_func(self,reconfiguration,solution,solution_idx):
        config = Configuration(param=self.param,varFlag=solution)
        pf = PowerFlow(config=config)
        r1 = pf.run1Config_WithObjective()
        fitness = -r1['Objective']
        return fitness
    def run(self,lineOff0=[],shuntOff0=[]):
        nIter = int(self.gaparam['nIter'])
        print('Running Genetic Algorithm nIter=%i'%nIter)
        param = Parameter(self.fi)
        rs = [[],[time.ctime(),'GeneticAlgorithm init_pos lineOff0=%s shuntOff0=%s'%(str(lineOff0),str(shuntOff0))],['iter','Objective','DeltaA','RateMax[%]','Umax[pu]','Umin[pu]','LineOff','ShuntOff'] ]
        r0 = {'Objective':math.inf,'LineOff':lineOff0,'ShuntOff':shuntOff0}
        gene_space = [tuple([0,1]) for _ in range(param.nVar)]
        start = time.time()
        reconfiguration = pygad.GA(
                        num_generations=nIter,
                        num_parents_mating=int(self.gaparam['num_parents_mating']),
                        fitness_func=self.fitness_func,
                        sol_per_pop=int(self.gaparam['sol_per_pop']),
                        mutation_num_genes=int(self.gaparam['mutation_num_genes']),
                        num_genes=param.nVar,gene_space=gene_space,
                        gene_type=int,keep_elitism=1,
                        crossover_probability=self.gaparam['crossover_probability'],
                        mutation_probability=self.gaparam['mutation_probability'],)
                            #parallel_processing=4 ,mutation_num_genes=10)
        reconfiguration.run()

        # Retrieve the best solution and its fitness value
        best_solution = reconfiguration.best_solution()

        # Print the best solution and its fitness value
        print("Best Solution: ", best_solution)
        #
        config = Configuration(param=param,varFlag=best_solution[0])
        lineOff = config.lineOff
        shuntOff = config.shuntOff
        pf = PowerFlow(config=config)
        res = pf.run1Config_WithObjective(fo=self.fo)
        s1 = ['GA','nIter',nIter,str(lineOff),str(shuntOff),str(best_solution[1]),res['Objective']]
        add2CSV(self.fo,[s1],',')
        pf = PowerFlow(config=config)
        res = pf.run1Config_WithObjective(self.fo)
        end = time.time()
        print(end-start)
        print(s1)
        #
        reconfiguration.plot_fitness()
        #
        return 
if __name__ == '__main__':
    
    ARGVS.fi = 'tromvia200percentSmax.xlsx'
    #ARGVS.fi = 'Inputs12_2-low2.xlsx'
    
    #ARGVS.fi = 'Inputs190month.xlsx'
    #ARGVS.fi = 'Inputs102.xlsx'
    
    #

    #ARGVS.fi = 'Inputs12_2-low.xlsx'
    lineOff0 = [] # no init
    shuntOff0 = []
    ARGVS.fo = 'res\\130withoutdg.csv'
    #PSO().run(1e4,lineOff0,shuntOff0)
    
    GA(ARGVS.fi,ARGVS.fo).run(lineOff0,shuntOff0)
    
    
    
