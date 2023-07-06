_author__    = "Cao Anh Quoc Hung ft Dr. Pham Quang Phuong"
__copyright__ = "Copyright 2023"
__license__   = "All rights reserved"
__email__     = "phuong.phamquang@hust.edu.vn"
__status__    = "Released"
__version__   = "1.1.5"
"""
about: ....
"""
import os,time,math
import openpyxl,csv
import argparse
import numpy as np


PARSER_INPUTS = argparse.ArgumentParser(epilog= "")
PARSER_INPUTS.usage = 'Distribution network analysis Tools'
PARSER_INPUTS.add_argument('-fi' , help = '*(str) Input file path' , default = '',type=str,metavar='')
PARSER_INPUTS.add_argument('-fo' , help = '*(str) Output file path', default = '',type=str,metavar='')
ARGVS = PARSER_INPUTS.parse_known_args()[0]
#
RATEC = 100/math.sqrt(3)

#
def toString(v,nRound=5):
    """ convert object/value to String """
    if v is None:
        return 'None'
    t = type(v)
    if t==str:
        if "'" in v:
            return ('"'+v+'"').replace('\n',' ')
        return ("'"+v+"'").replace('\n',' ')
    if t==int:
        return str(v)
    if t==float:
        if v>1.0:
            s1 = str(round(v,nRound))
            return s1[:-2] if s1.endswith('.0') else s1
        elif abs(v)<1e-8:
            return '0'
        s1 ='%.'+str(nRound)+'g'
        return s1 % v
    if t==complex:
        if v.imag>=0:
            return '('+ toString(v.real,nRound)+' +' + toString(v.imag,nRound)+'j)'
        return '('+ toString(v.real,nRound) +' '+ toString(v.imag,nRound)+'j)'
    try:
        return v.toString()
    except:
        pass
    if t in {list,tuple,set}:
        s1=''
        for v1 in v:
            s1+=toString(v1,nRound)+','
        if v:
            s1 = s1[:-1]
        if t==list:
            return '['+s1+']'
        elif t==tuple:
            return '('+s1+')'
        else:
            return '{'+s1+'}'
    if t==dict:
        s1=''
        for k1,v1 in v.items():
            s1+=toString(k1)+':'
            s1+=toString(v1,nRound)+','
        if s1:
            s1 = s1[:-1]
        return '{'+s1+'}'
    return str(v)
#
def __checkLoop__(busHnd,bus,br):
    setBusChecked = set()
    setBrChecked = set()
    for o1 in busHnd:
        if o1 not in setBusChecked:
            setBusChecked.add(o1)
            tb1 = {o1}
            #
            for i in range(20000):
                if i==19999:
                    raise Exception('Error in checkLoop()')
                tb2 = set()
                for b1 in tb1:
                    for l1 in bus[b1]:
                        if l1 not in setBrChecked:
                            setBrChecked.add(l1)
                            for bi in br[l1]:
                                if bi!=b1:
                                    if bi in setBusChecked or bi in tb2:
                                        return bi
                                    tb2.add(bi)
                if len(tb2)==0:
                    break #ok finish no loop for this group
                setBusChecked.update(tb2)
                tb1=tb2.copy()
    return None
#
def __findBusConnected__(bi1,bset,lset):
    ## find all bus connected to Bus b1
    res = {bi1}
    ba = {bi1}
    while True:
        ba2 = set()
        for b1 in ba:
            for li in bset[b1]:
                for bi in lset[li]:
                    if bi not in res:
                        ba2.add(bi)
        if ba2:
            res.update(ba2)
            ba=ba2
        else:
            break
    return res
#
def __getLineISL__(busC0):
    lineISL = set() # line can not be off => ISLAND
    busC = busC0.copy()
    while True:
        n1 = len(lineISL)
        for k,v in busC.items():
            if len(v)==1:
                lineISL.update(v)
        if n1==len(lineISL):
            break
        busc1 = dict()
        for k,v in busC.items():
            if len(v)!=1:
                busc1[k]=v-lineISL
        busC = busc1.copy()
    return lineISL,busC
#
def add2CSV(nameFile,ares,delim):
    """
    append array String to a file CSV
    """
    pathdir = os.path.split(os.path.abspath(nameFile))[0]
    try:
        os.mkdir(pathdir)
    except:
        pass
    #
    if not os.path.isfile(nameFile):
        with open(nameFile, mode='w') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n")
            for a1 in ares:
                ew.writerow(a1)
            f.close()
    else:
        with open(nameFile, mode='a') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
            for a1 in ares:
                ew.writerow(a1)
#
class Parameter:
    def __init__(self,fi):
        self.tcheck = 0
        wbInput = openpyxl.load_workbook(os.path.abspath(fi),data_only=True)
        loadProfile = self.__readInput1Sheet__(wbInput,'LOADPROFILE')
        genProfile = self.__readInput1Sheet__(wbInput,'GENPROFILE')
        busa = self.__readInput1Sheet__(wbInput,'BUS')
        linea = self.__readInput1Sheet__(wbInput,'LINE')
        self.setting = self.__readSetting__(wbInput)
        #
        self.iterMax = int(self.setting['option_PF'][0])
        self.epsilon = self.setting['option_PF'][1]
        self.lineOff = None
        # print(self.setting)
        # BUS[NO] =[ kV,PLOAD[kw],QLOAD[kvar],code ]
        self.BUS = {}
        self.busSlack = []
        self.BUSbs = {} #shunt
        self.Qsh = {}
        for i in range(len(busa['NO'])):
            if busa['FLAG'][i]:
                n1 = busa['NO'][i]
                kv = busa['kV'][i]
                p1 = busa['PLOAD[kw]'][i]/1000
                q1 = busa['QLOAD[kvar]'][i]/1000
                qsh = busa['Qshunt[kvar]'][i]/1000
                if abs(qsh)>1e-6:
                    self.BUSbs[n1] = qsh
                #
                c1 = busa['CODE'][i]
                if c1==None:
                    c1=1
                if c1 in {2,3}:
                    self.busSlack.append(n1)
                #
                self.BUS[n1] = [kv,p1,q1,c1]
        self.nSlack = len(self.busSlack)
        self.setSlack = set(self.busSlack)
        #
        self.Ubase = self.BUS[self.busSlack[0]][0]
        self.Ubase2 = self.Ubase*self.Ubase
        # update B shunt at bus
        for k1,v1 in self.BUSbs.items():
            self.BUSbs[k1] = v1/self.Ubase2# q =u*u*b
        #
        self.profileID = [int(i) for i in loadProfile['time\\NOBUS']]
        # LOAD PROFILE convert to MVA
        self.loadProfile = dict()
        self.loadAll = dict()
        for ii in range(len(self.profileID)):
            k = loadProfile['time\\NOBUS'][ii]
            v1 = dict()
            self.loadAll[k] = 0
            for k1 in loadProfile.keys():
                if k1!='time\\NOBUS':
                    n1 = int(k1)
                    v1[n1] = loadProfile[k1][ii] * complex(self.BUS[n1][1],self.BUS[n1][2])
                    self.loadAll[k]+=v1[n1]
            self.loadProfile[k] = v1
        # GENPROFILE convert to kV
        self.genProfile = dict()
        for ii in range(len(self.profileID)):
            k = genProfile['time\\NOBUS'][ii]
            v1 = dict()
            for k1 in genProfile.keys():
                if k1!='time\\NOBUS':
                    n1 = int(k1)
                    v1[n1] = genProfile[k1][ii] * self.Ubase
            self.genProfile[k]=v1
        # LINE[NO] = [FROMBUS,TOBUS,RX(Ohm),B/2(Siemens),RATE ]
        self.LINE = {}
        self.LINEb = {} # b of Line
        for i in range(len(linea['NO'])):
            if linea['FLAG'][i]:
                n1= linea['NO'][i]
                fr = linea['FROMBUS'][i]
                to = linea['TOBUS'][i]
                r = linea['R(Ohm)'][i]
                x = linea['X(Ohm)'][i]
                r1 = linea['RATEA[A]'][i]/1000 #kA
                self.LINE[n1] = [fr,to,complex(r,x),r1]
                #
                if linea['B(microSiemens)'][i]>1e-2:
                    self.LINEb[n1] = linea['B(microSiemens)'][i]*1e-6/2
        self.setLineHndAll = set(self.LINE.keys())
        #
        self.setBusHnd = set(self.BUS.keys())
        self.lstBusHnd = busa['NO']
        self.lstLineHnd= linea['NO']
        # list cac line co the dong mo
        self.lineFLAG3 = []
        for i in range(len(linea['NO'])):
            if linea['FLAG3'][i]:
                self.lineFLAG3.append(linea['NO'][i])
        #
        self.shuntFLAG3 = []
        if 'FLAG3' in busa.keys():
            for i in range(len(busa['NO'])):
                if busa['FLAG3'][i]:
                    self.shuntFLAG3.append(busa['NO'][i])
        #
        self.nVar = len(self.lineFLAG3) + len(self.shuntFLAG3)
        self.nL = len(self.lineFLAG3)
        #
        self.BUSC = dict() #connect of BUS
        for b1 in self.setBusHnd:
            self.BUSC[b1] = set()
        #
        for k,v in self.LINE.items():
            self.BUSC[v[0]].add(k)
            self.BUSC[v[1]].add(k)
        #
        self.t0 = time.time()
        #print(self.BUSC)
        #print(self.lineISL) # line ko the off, off=>island
        #assume that only take circumstances where there is no island bus to calculate power flow
        self.nBus = len(self.setBusHnd)
        #print('busISL',self.busISL)   # bus con lai sau khi da bo line island, dung de check loop
    def __readInput1Sheet__(self,wbInput,sh1):
        ws = wbInput[sh1]
        res = {}
        # dem so dong data
        for i in range(2,20000):
            if ws.cell(i,1).value==None:
                k=i
                break
        #
        for i in range(1,20000):
            v1 = ws.cell(2,i).value
            if v1==None:
                return res
            va = []
            #
            for i1 in range(3,k):
                va.append(ws.cell(i1,i).value)
            res[str(v1)]=va
        return res
    #
    def __readSetting__(self,wbInput):
        ws = wbInput['SETTING']
        k = 0
        res = {}
        while True:
            k+=1
            s1= ws.cell(k,1).value
            if type(s1)==str and s1.replace(' ','')=='##BRANCHING':
                for j in range(1,100):
                    s2 = str(ws.cell(k+1,j).value).strip()
                    if s2=='None':
                        break
                    sa = str(ws.cell(k+2,j).value).split(',')
                    if len(sa)==1:
                        try:
                            res[s2] = float(sa[0])
                        except:
                            res[s2] = sa[0]
                    else:
                        res[s2] = [float(si) for si in sa]
                break
        return res
    #

class Configuration:
    def __init__(self,param: Parameter,lineOff=[],shuntOff=[],varFlag=None):
        self.param = param
        if varFlag is not None:
            if len(varFlag)!=self.param.nVar:
                raise Exception('Error size of varFlag')
            lineOff = self.getLineOff(varFlag[:self.param.nL])
            shuntOff = self.getShuntOff(varFlag[self.param.nL:])
        self.setBusHnd =self.param.setBusHnd
        self.setLineHnd = self.param.setLineHndAll - set(lineOff)
        self.setLinebHnd = self.param.LINEb.keys() - set(lineOff)
        self.shuntOff = set(shuntOff)
        self.lineOff = set(lineOff)
        self.lineSureISL,busc1 = self.__getLineISL__(self.param.BUSC)
        self.bus0ISL = set(busc1.keys())
        self.busISL = self.setBusHnd - self.bus0ISL
        #
        #
        self.lineC = {k:self.param.LINE[k][:2] for k in self.setLineHnd}
        self.busC = {b1:set() for b1 in self.setBusHnd}
        for k,v in self.lineC.items():
            self.busC[v[0]].add(k)
            self.busC[v[1]].add(k)
        #
    #
    def getLineFlag3(self):
        """ cac Branch co the dong mo """
        return self.param.lineFLAG3
    #
    def getShuntFlag3(self):
        """ cac Shunt co the dong mo """
        return self.param.shuntFLAG3
    #
    def getLineOff(self,lineFlag): # 0: inservice, 1 off service
        lineOff = []
        for i in range(len(self.param.lineFLAG3)):
            if lineFlag[i]:
                lineOff.append(self.param.lineFLAG3[i])
        return lineOff
    #
    def getShuntOff(self,shuntFlag): # 0: inservice, 1 off service
        shuntOff = []
        for i in range(len(self.param.shuntFLAG3)):
            if shuntFlag[i]:
                shuntOff.append(self.param.shuntFLAG3[i])
        return shuntOff
    #
    def getVarFlag(self,lineOff,shuntOff):
        varFlag = [0]*self.param.nVar
        for i in range(self.param.nL):
            if self.param.lineFLAG3[i] in lineOff:
                varFlag[i] = 1
        for i in range(self.param.nVar-self.param.nL):
            if self.param.shuntFLAG3[i] in shuntOff:
                varFlag[self.param.nL+i] = 1
        return varFlag
    def __checkLoopIsland__(self,lineOff):
        # check island/loop multi slack ----------------------------------------
        if lineOff.intersection(self.lineSureISL):
            return 'ISLAND'
        #
        self.setLineHnd = self.param.setLineHndAll - lineOff
        
        r11 = self.setBusHnd.copy()
        self.busGroup = []# cac bus tuong ung o cac slack khac nhau
        for bs1 in self.param.busSlack:
            r1 = self.__findBusConnected__(bs1,self.busC,self.lineC)
            if len(r1.intersection(self.param.setSlack))>1:
                return 'LOOP MULTI SLACK'
            #
            self.busGroup.append(r1)
            r11.difference_update(r1)
        #
        if r11:
            return 'ISLAND'
        #
        # LOOP
        if self.__checkLoop__(self.bus0ISL,self.busC,self.lineC):
            return 'LOOP'
        #
        return ''
        
    def __findBusConnected__(self,bi1,bset,lset):
        ## find all bus connected to Bus b1
        res = {bi1}
        ba = {bi1}
        while True:
            ba2 = set()
            for b1 in ba:
                for li in bset[b1]:
                    for bi in lset[li]:
                        if bi not in res:
                            ba2.add(bi)
            if ba2:
                res.update(ba2)
                ba=ba2
            else:
                break
        return res
    def __checkLoop__(self,busHnd,bus,br):
        setBusChecked = set()
        setBrChecked = set()
        for o1 in busHnd:
            if o1 not in setBusChecked:
                setBusChecked.add(o1)
                tb1 = {o1}
                #
                for i in range(20000):
                    if i==19999:
                        raise Exception('Error in checkLoop()')
                    tb2 = set()
                    for b1 in tb1:
                        for l1 in bus[b1]:
                            if l1 not in setBrChecked:
                                setBrChecked.add(l1)
                                for bi in br[l1]:
                                    if bi!=b1:
                                        if bi in setBusChecked or bi in tb2:
                                            return bi
                                        tb2.add(bi)
                    if len(tb2)==0:
                        break #ok finish no loop for this group
                    setBusChecked.update(tb2)
                    tb1=tb2.copy()
        return None
    def __getLineISL__(self,busC0):
        lineISL = set() # line can not be off => ISLAND
        busC = busC0.copy()
        while True:
            n1 = len(lineISL)
            for k,v in busC.items():
                if len(v)==1:
                    lineISL.update(v)
            if n1==len(lineISL):
                break
            busc1 = dict()
            for k,v in busC.items():
                if len(v)!=1:
                    busc1[k]=v-lineISL
            busC = busc1.copy()
        return lineISL,busC

class RunMethod:
    def __init__(self,param: Parameter,config:Configuration):
        self.param = param
        self.config = config
        self.lineOff = self.config.lineOff
        self.shuntOff = self.config.shuntOff
    def __initcsv__(self,fo):
        add2CSV(fo,[[],[time.ctime()],['PF 1Profile','lineOff',str(list(self.lineOff)),'shuntOff',str(list(self.shuntOff))]],',')
        #
        rB = [[],['BUS/Profile']]
        rB[1].extend([bi for bi in self.param.lstBusHnd])
        #
        rL = [[],['LINE/Profile']]
        rL[1].extend([bi for bi in self.param.lstLineHnd])
        #
        rG = [[],['GEN/Profile']]
        for bi in self.param.busSlack:
            rG[1].append(str(bi)+'_P')
            rG[1].append(str(bi)+'_Q')
            rG[1].append(str(bi)+'_cosPhi')
        return rB,rG,rL
    def __update1profile__(self,pi,rB,rL,rG,sa1,va1,dia1):
        #sa1,dia1,va1 = dict(),dict(),dict() # for 1 profile
        rb1 = [pi]
        rl1 = [pi]
        rg1 = [pi]
        for bi1 in self.param.lstBusHnd:
            rb1.append(toString(abs(va1[bi1])/self.param.Ubase))
        #
        for bri in self.param.lstLineHnd:
            try:
                r1 = abs(dia1[bri])/self.param.LINE[bri][3]*RATEC
                rl1.append( toString(r1,2) )
            except:
                rl1.append('0')
        #
        for bs1 in self.param.busSlack:
            rg1.append(toString(sa1[bs1].real))
            rg1.append(toString(sa1[bs1].imag))
            if sa1[bs1].imag>=0:
                rg1.append(toString(sa1[bs1].real/abs(sa1[bs1]),3))
            else:
                rg1.append(toString(-sa1[bs1].real/abs(sa1[bs1]),3))
        #
        rB.append(rb1)
        rL.append(rl1)
        rG.append(rg1)
        return rB,rL,rG
    
    def __update_result__(self,res,va,ra,cosP,cosN):
        res['Umax[pu]'] = max(va)/self.param.Ubase
        res['Umin[pu]'] = min(va)/self.param.Ubase
        res['RateMax[%]'] = max(ra)
        res['cosP'] = min(cosP)
        res['cosN'] = max(cosN)
        return res

    def __export_profiles__(self,fo,rB,rL,rG,res):
        rB.append(['','Umax[pu]',toString(res['Umax[pu]']),'Umin[pu]',toString(res['Umin[pu]']) ])
        add2CSV(fo,rB,',')
        #
        rL.append(['','RateMax[%]',toString(res['RateMax[%]'],2)])
        add2CSV(fo,rL,',')
        #
        rG.append(['','cosPmin',toString(res['cosP'],3),'cosNMax',toString(res['cosN'],3)])
        add2CSV(fo,rG,',')
        return    

class YbusMatrix:
    def __init__(self,param:Parameter,config:Configuration):
        self.param = param
        self.config = config
        self.setLineHnd = config.setLineHnd
        self.setLinebHnd = config.setLinebHnd
        self.shuntOff = config.shuntOff

    def __get_sparse_Ybus__(self):
        sparse_ybus = dict()
        #initialize branch admittance
        y = {k:(1+0j) for k in self.setLineHnd}
        #formation of the off diagonal elements
        for k,v in self.config.lineC.items():
            y[k] = y[k]/self.param.LINE[k][2]
            sparse_ybus[frozenset({v[0],v[1]})] = -y[k]   
        #calculate yline with lineb: 
        for lbi in self.setLinebHnd:
            y[lbi] += self.param.LINEb[lbi]*1j
        # Shunt 
        for k1,v1 in self.param.BUSbs.items():
            if k1 not in self.shuntOff:
                sparse_ybus[k1] = v1*1j
        #formation of the diagonal elements
        for bi in self.config.busC.keys():
            for li in self.config.busC[bi]:
                if bi not in sparse_ybus:
                    sparse_ybus[bi] =  y[li]
                else:
                    sparse_ybus[bi] +=  y[li]
        # case if bus slack does not connect to any bus
        for bsi in self.param.busSlack:
            if not self.config.busC[bsi]:
                sparse_ybus[bsi] = 0
        return sparse_ybus
    #
    def __getYbus__(self):
        #initialize Ybus
        Ybus = []
        for _ in range(self.param.nBus):
            row = [0] * self.param.nBus
            Ybus.append(row)
        #initialize branch admittance
        y = {k:(1+0j) for k in self.config.setLineHnd}
        #formation of the off diagonal elements
        for k,v in self.config.lineC.items():
                y[k] = y[k]/self.param.LINE[k][2]
                Ybus[v[0]-1][v[1]-1] = -y[k]
                Ybus[v[1]-1][v[0]-1] = Ybus[v[0]-1][v[1]-1]
        #calculate yline with lineb: 
        for lbi in self.setLinebHnd:
            y[lbi] += self.param.LINEb[lbi]*1j
        #formation of the diagonal elements
        for bi in self.config.busC.keys():
            for li in self.config.busC[bi]:
                Ybus[bi-1][bi-1] +=  y[li]
        # Shunt 
        for k1,v1 in self.param.BUSbs.items():
            if k1 not in self.shuntOff:
                Ybus[k1-1][k1-1] += v1*1j
        return Ybus

class GaussSeidel(RunMethod):
    def __init__(self,param:Parameter,config:Configuration):
        super().__init__(param=param,config=config)
        self.accel = self.param.setting['accel']
    def __run1config__(self,fo=''):
        # ready to run Gauss Seidel
        Ybus = YbusMatrix(self.param,self.config).__get_sparse_Ybus__()
        # initialize result
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
        #
        if fo:
            rB,rG,rL = super().__initcsv__(fo)
        va,ra,cosP,cosN = [],[],[1],[-1]
        for pi in self.param.profileID:
            P = {bi:(-self.param.loadProfile[pi][bi]).real for bi in self.config.setBusHnd}    #+self.Pgen
            Q = {bi:(-self.param.loadProfile[pi][bi]).imag for bi in self.config.setBusHnd}    #+self.Qgen
            DP = {bi:0 for bi in self.param.setBusHnd}
            DQ = {bi:0 for bi in self.param.setBusHnd}
            sbus = {bi:(P[bi] + Q[bi]*1j) for bi in self.config.setBusHnd}
            vbus = {bi:complex(self.param.Ubase,0) for bi in self.config.setBusHnd}
            # initialize voltage correction
            Vc = {bi:complex(0,0) for bi in self.config.setBusHnd}
            # update Vm of slack buses
            for bs in self.param.setSlack:
                vbus[bs] = complex(self.param.genProfile[pi][bs],0) 
            Vm = {bi:abs(vbus[bi])for bi in self.config.setBusHnd}
            
            for ii in range(self.param.iterMax+1):
                for b1 in self.config.setBusHnd:
                    YV = 0 + 1j * 0
                    for li in self.config.busC[b1]:
                        for b2 in self.config.lineC[li]:
                            if b1 == b2:
                                pass
                            else:
                                line = frozenset({b1,b2})
                                YV += Ybus[line] * vbus[b2]
                    Sc = np.conj(vbus[b1]) * (Ybus[b1] * vbus[b1] + YV)
                    Sc = np.conj(Sc)
                    DP[b1] = P[b1] - np.real(Sc)
                    DQ[b1] = Q[b1] - np.imag(Sc)
                    if self.param.BUS[b1][3] == 3:
                        sbus[b1] = Sc
                        P[b1] = np.real(Sc)
                        Q[b1] = np.imag(Sc)
                        DP[b1] = 0
                        DQ[b1] = 0
                        Vc[b1] = vbus[b1]
                    elif self.param.BUS[b1][3] == 2:
                        Q[b1] = np.imag(Sc)
                        sbus[b1] = P[b1] + 1j * Q[b1]
                    if self.param.BUS[b1][3] != 3:
                        Vc[b1] = (np.conj(sbus[b1]) / np.conj(vbus[b1]) - YV) / Ybus[b1]
                    else:
                        pass
                    if self.param.BUS[b1][3] == 1:
                        vbus[b1] = vbus[b1] + self.accel * (Vc[b1] - vbus[b1])
                    elif self.param.BUS[b1][3] == 2:
                        VcI = np.imag(Vc[b1])
                        VcR = np.sqrt(Vm[b1] ** 2 - VcI ** 2)
                        Vc[b1] = VcR + 1j * VcI
                        vbus[b1] = vbus[b1] + self.accel * (Vc[b1] - vbus[b1])
                maxerror = abs(DP[b1])
                for bi in self.config.setBusHnd:
                    if abs(DP[bi]) > maxerror:
                        maxerror = abs(DP[bi])
                    if abs(DQ[bi]) > maxerror:
                        maxerror = abs(DQ[bi])
                if maxerror < self.param.epsilon:
                    break
                if ii==self.param.iterMax:
                    return {'FLAG':'DIVERGENCE'}
            # finish GS
            # store all voltage magnitude value in all profile
            Vm = {bi:abs(vbus[bi]) for bi in self.config.setBusHnd}
            va.extend(Vm.values())
            for bi in self.param.busSlack:
                sbus[bi] += self.param.loadProfile[pi][bi]
            """
            for bi in self.busPV:
                # update Q load in PV bus
                sbus[bi] += self.loadProfile[pi][bi].imag * 1j 
            """
            #intialize s loss total
            slt = 0
            Il = dict()
            for li,bi in self.config.lineC.items():
                line = frozenset({bi[0],bi[1]})
                Ib1 = (vbus[bi[0]]-vbus[bi[1]])*(-Ybus[line])
                Ib2 = (vbus[bi[1]]-vbus[bi[0]])*(-Ybus[line])
                #
                if abs(Ib1) >= abs(Ib2):
                    Il[li] = abs(Ib1)
                else: 
                    Il[li] = abs(Ib2)   
                rate = ((Il[li])/self.param.LINE[li][3])*RATEC
                ra.append(rate)
                Snk = vbus[bi[0]]*np.conj(Ib1)
                Skn = vbus[bi[1]]*np.conj(Ib2)
                slt += Snk +Skn
            if fo:
                sa1,va1,dia1 = dict(),dict(),dict()# for 1 profile
                va1.update(vbus)
                dia1.update(Il)
                sa1.update(sbus)
                rB,rL,rG =super().__update1profile__(pi,rB,rL,rG,sa1,va1,dia1)
                
            res['DeltaA'] += slt.real
        res = super().__update_result__(res,va,ra,cosP,cosN)
        if fo:
            super().__export_profiles__(fo,rB,rL,rG,res)
        return res

class SparseNewtonRaphson(RunMethod):
    def __init__(self,param:Parameter,config:Configuration):
        super().__init__(param=param,config=config)
    def __run1config__(self, fo=''):
        # ready to run Newton raphson
        Ybus = YbusMatrix(self.param,self.config).__get_sparse_Ybus__()
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
        #
        if fo:
            rB,rG,rL = super().__initcsv__(fo)
        Ym, theta = dict(),dict()
        for k,v in Ybus.items():
            Ym[k] = abs(v)
            theta[k] = np.angle(v,deg=False)
        #
        countSlack = 0
        countPV = 0
        slackCounted = []
        PVcounted = []
        for bi in self.config.setBusHnd:
            if self.param.BUS[bi][3] == 3:
                countSlack+=1 
            elif self.param.BUS[bi][3] == 2:
                countPV += 1 
            slackCounted.append(countSlack)
            PVcounted.append(countPV)
        no_jacobi_equation = 2 * self.param.nBus - countPV - 2 * self.param.nSlack         
        #
        DC = np.zeros(no_jacobi_equation) 
        #
        va,ra,cosP,cosN = [],[],[1],[-1]
        for pi in self.param.profileID:
            # initialize voltage magnitude of all buses
            # iterate in setBUShnd in case some buses are off 
            Vm = {bi:float(self.param.Ubase) for bi in self.config.setBusHnd}
            # initialize voltage angle 
            delta = {bi:0 for bi in self.config.setBusHnd}
            P = {bi:(-self.param.loadProfile[pi][bi]).real for bi in self.config.setBusHnd}    #+self.Pgen
            Q = {bi:(-self.param.loadProfile[pi][bi]).imag for bi in self.config.setBusHnd}    #+self.Qgen
            # update Vm of slack buses
            for bs in self.param.setSlack:
                Vm[bs] = self.param.genProfile[pi][bs]
            for ii in range(self.param.iterMax+1):
                # Initialize Jacobian Matrix
                A = np.zeros((no_jacobi_equation,no_jacobi_equation))
                for b1 in self.config.setBusHnd:

                    J1_row_offdiag_idx = J2_row_offdiag_idx = int((b1 - slackCounted[b1-1])-1)
                    J1_diag_idx = J2_row_diag_idx = J3_col_diag_idx = J1_row_offdiag_idx 

                    J3_row_offdiag_idx = J4_row_offdiag_idx = int((self.param.nBus+b1-slackCounted[b1-1]-PVcounted[b1-1]-self.param.nSlack)-1)
                    J4_diag_idx = J2_col_diag_idx = J3_row_diag_idx = J3_row_offdiag_idx 

                    J11 = 0
                    J22 = 0
                    J33 = 0
                    J44 = 0
                    #
                    for li in self.config.busC[b1]:
                        for b2 in self.config.lineC[li]:
                            if b1 == b2:
                                pass
                            else:
                                line = frozenset({b1,b2})
                                # diagonal elements of J1
                                J11 += Vm[b1] * Vm[b2] * Ym[line] * math.sin((theta[line] - delta[b1] + delta[b2]).real)
                                # diagonal elements of J3          
                                J33 += Vm[b1] * Vm[b2] * Ym[line] * math.cos((theta[line] - delta[b1] + delta[b2]).real)
                                if self.param.BUS[b1][3] != 3:
                                    #slackbus doesn't exist in J2 and J4
                                    #diagonal elements of J2
                                    J22 += Vm[b2] * Ym[line] * math.cos((theta[line] - delta[b1] + delta[b2]).real)
                                    # diagonal elements of J4               
                                    J44 += Vm[b2] * Ym[line] * math.sin((theta[line] - delta[b1] + delta[b2]).real)
                                if self.param.BUS[b1][3] != 3 and self.param.BUS[b2][3] != 3:
                                    J1_col_offdiag_idx = J3_col_offdiag_idx = int(b2 - slackCounted[b2-1] - 1)
                                    J2_col_offdiag_idx = J4_col_offdiag_idx = int(self.param.nBus + b2 - PVcounted[b2-1] - slackCounted[b2-1] - self.param.nSlack - 1)
                                    # off diagonal elements of J1
                                    A[J1_row_offdiag_idx][J1_col_offdiag_idx] = -Vm[b1] * Vm[b2] * Ym[line] * math.sin((theta[line] - delta[b1] + delta[b2]).real)
                                
                                    if self.param.BUS[b2][3] == 1:
                                        # off diagonal elements of J2
                                        A[J2_row_offdiag_idx][J2_col_offdiag_idx] = Vm[b1] * Ym[line] * math.cos((theta[line] - delta[b1] + delta[b2]).real)
                                
                                    if self.param.BUS[b1][3] == 1:
                                        # off diagonal elements of J3
                                        A[J3_row_offdiag_idx][J3_col_offdiag_idx] = -Vm[b1] * Vm[b2] * Ym[line] * math.cos((theta[line] - delta[b1] + delta[b2]).real)
                                
                                    if self.param.BUS[b1][3] == 1 and self.param.BUS[b2][3] == 1:
                                        # off diagonal elements of J4
                                        A[J4_row_offdiag_idx][J4_col_offdiag_idx] = -Vm[b1] * Ym[line] * math.sin((theta[line] - delta[b1] + delta[b2]).real)
                    #   
                    Pk = Vm[b1]**2 * Ym[b1] * math.cos((theta[b1])) + J33
                    Qk = -Vm[b1]**2 * Ym[b1] * math.sin((theta[b1])) - J11
                    if self.param.BUS[b1][3] == 3:
                        # swing bus
                        P[b1] = Pk
                        Q[b1] = Qk
                    if self.param.BUS[b1][3] == 2:
                        Q[b1] = Qk
                        """
                        if Qmax[b1-1] != 0:
                            Qgc = Q[b1-1] + Qd[n-1] - Qsh[n-1]
                            if ii <= 7:                   #Between the 2th & 6th iterations
                                if ii > 2:                #the Mvar of generator buses are
                                    if Qgc < Qmin[n-1]:   #tested. If not within limits Vm(n)
                                        Vm[n-1] += 0.01   #is changed in steps of 0.01 pu to
                                    elif Qgc > Qmax[n-1]: #bring the generator Mvar within
                                        Vm[n-1] -= 0.01   #the specified limits.            
                        """
                    if self.param.BUS[b1][3] != 3:
                        # diagonal elements of J1
                        A[J1_diag_idx][J1_diag_idx] = J11        
                        DC[J1_diag_idx] = P[b1] - Pk
                    if self.param.BUS[b1][3] == 1:
                        # diagonal elements of J2
                        A[J2_row_diag_idx][J2_col_diag_idx] = 2 * Vm[b1] * Ym[b1] * math.cos(theta[b1]) + J22    
                        # diagonal elements of J3
                        A[J3_row_diag_idx][J3_col_diag_idx] = J33
                        # diagonal elements of J4         
                        A[J4_diag_idx][J4_diag_idx] = -2 * Vm[b1] * Ym[b1] * math.sin(theta[b1]) - J44   
                        DC[J4_diag_idx] = Q[b1] - Qk
                #matrix A left division Matrix DC
                DX = np.linalg.solve(A, DC.T)
                for bi in self.config.setBusHnd:
                    del_update_DXidx = int(bi - slackCounted[bi-1]-1)
                    Vm_update_DXidx = int(self.param.nBus + bi - PVcounted[bi-1] - slackCounted[bi-1] - self.param.nSlack-1)
                    if self.param.BUS[bi][3] != 3:
                        delta[bi] +=  DX[del_update_DXidx]
                    if self.param.BUS[bi][3] == 1:
                        Vm[bi] += DX[Vm_update_DXidx]
                maxerror = max(abs(DC))
                if maxerror < self.param.epsilon:
                    break
                if ii==self.param.iterMax:
                    return {'FLAG':'DIVERGENCE'}
            # finish NR
            # store all voltage magnitude value in all profile
            va.extend(Vm.values())
            vbus = {bi:(Vm[bi]* np.cos(delta[bi])+Vm[bi]*1j*np.sin(delta[bi])) for bi in self.config.setBusHnd }
            sbus = {bi:(P[bi] + Q[bi]*1j) for bi in self.config.setBusHnd}
            for bi in self.param.busSlack:
                sbus[bi] += self.param.loadProfile[pi][bi]
            """
            for bi in self.busPV:
                # update Q load in PV bus
                sbus[bi] += self.loadProfile[pi][bi].imag * 1j 
            """
            slt = 0
            Il = dict()
            for li,bi in self.config.lineC.items():
                line = frozenset({bi[0],bi[1]})
                Ib1 = (vbus[bi[0]]-vbus[bi[1]])*(-Ybus[line])
                Ib2 = (vbus[bi[1]]-vbus[bi[0]])*(-Ybus[line])
                #
                if abs(Ib1) >= abs(Ib2):
                    Il[li] = abs(Ib1)
                else: 
                    Il[li] = abs(Ib2)   
                rate = ((Il[li])/self.param.LINE[li][3])*RATEC
                ra.append(rate)
                
                Snk = vbus[bi[0]]*np.conj(Ib1)
                Skn = vbus[bi[1]]*np.conj(Ib2)
                slt += Snk +Skn 
            if fo:
                sa1,va1,dia1 = dict(),dict(),dict()# for 1 profile
                va1.update(vbus)
                dia1.update(Il)
                sa1.update(sbus)
                rB,rL,rG =super().__update1profile__(pi,rB,rL,rG,sa1,va1,dia1)
            res['DeltaA'] += slt.real
        res = super().__update_result__(res,va,ra,cosP,cosN)
        if fo:
            super().__export_profiles__(fo,rB,rL,rG,res)
        return res
    
class NewtonRaphson(RunMethod):
    def __init__(self,param:Parameter,config:Configuration):
        super().__init__(param=param,config=config)
    def __run1config__(self, fo=''):
        # ready to run Newton raphson

        Ybus = YbusMatrix(self.param,self.config).__getYbus__()
        
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
        #
        if fo:
            rB,rG,rL = super().__initcsv__(fo)
        #
        # return Y bus magnitude
        Ym = np.abs(Ybus)
        # return phase angle of Ybus        
        theta = np.angle(Ybus,deg=False)
        countSlack = 0
        countPV = 0
        slackCounted = []
        PVcounted = []
        for bi in self.param.BUS.keys():
            if self.param.BUS[bi][3] == 3:
                countSlack+=1 
            elif self.param.BUS[bi][3] == 2:
                countPV += 1 
            slackCounted.append(countSlack)
            PVcounted.append(countPV)
        no_jacobi_equation = 2 * self.param.nBus - countPV - 2 * self.param.nSlack
        #
        DC = np.zeros(no_jacobi_equation)
        #
        va,ra,cosP,cosN = [],[],[1],[-1]
        for pi in self.param.profileID:
            # initialize voltage magnitude of all buses
            Vm = [float(self.param.Ubase) for _ in self.config.setBusHnd]
            P = [(-self.param.loadProfile[pi][bi]).real for bi in self.config.setBusHnd]    #+self.Pgen
            Q = [(-self.param.loadProfile[pi][bi]).imag for bi in self.config.setBusHnd]    #+self.Qgen
            # update Vm of slack buses
            for bs in self.param.setSlack:
                Vm[bs-1] = self.param.genProfile[pi][bs]
            # initialize voltage angle 
            delta = [0 for _ in self.config.setBusHnd]
            sa1,dia1,va1 = dict(),dict(),dict()# for 1 profile
            for ii in range(self.param.iterMax+1):
                A = np.zeros((no_jacobi_equation,no_jacobi_equation))
                for b1 in self.config.setBusHnd:
                    J1_row_offdiag_idx = J2_row_offdiag_idx = int((b1 - slackCounted[b1-1])-1)
                    J1_diag_idx = J2_row_diag_idx = J3_col_diag_idx = J1_row_offdiag_idx 

                    J3_row_offdiag_idx = J4_row_offdiag_idx = int((self.param.nBus+b1-slackCounted[b1-1]-PVcounted[b1-1]-self.param.nSlack)-1)
                    J4_diag_idx = J2_col_diag_idx = J3_row_diag_idx = J3_row_offdiag_idx 

                    J11 = 0
                    J22 = 0
                    J33 = 0
                    J44 = 0
                    #
                    for li in self.config.busC[b1]:
                        for b2 in self.config.lineC[li]:
                            if b1 == b2:
                                pass
                            else:
                                # diagonal elements of J1
                                J11 += Vm[b1-1] * Vm[b2-1] * Ym[b1-1][b2-1] * math.sin((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                # diagonal elements of J3          
                                J33 += Vm[b1-1] * Vm[b2-1] * Ym[b1-1][b2-1] * math.cos((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                if self.param.BUS[b1][3] != 3:
                                    #slackbus doesn't exist in J2 and J4
                                    #diagonal elements of J2
                                    J22 += Vm[b2-1] * Ym[b1-1][b2-1] * math.cos((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                    #diagonal elements of J4               
                                    J44 += Vm[b2-1] * Ym[b1-1][b2-1] * math.sin((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)

                                if self.param.BUS[b1][3] != 3 and self.param.BUS[b2][3] != 3:
                                    J1_col_offdiag_idx = J3_col_offdiag_idx = int(b2 - slackCounted[b2-1] - 1)
                                    J2_col_offdiag_idx = J4_col_offdiag_idx = int(self.param.nBus + b2 - PVcounted[b2-1] - slackCounted[b2-1] - self.param.nSlack - 1)
                                    # off diagonal elements of J1
                                    A[J1_row_offdiag_idx][J1_col_offdiag_idx] = -Vm[b1-1] * Vm[b2-1] * Ym[b1-1][b2-1] * math.sin((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                
                                    if self.param.BUS[b2][3] == 1:
                                        # off diagonal elements of J2
                                        A[J2_row_offdiag_idx][J2_col_offdiag_idx] = Vm[b1-1] * Ym[b1-1][b2-1] * math.cos((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                
                                    if self.param.BUS[b1][3] == 1:
                                        # off diagonal elements of J3
                                        A[J3_row_offdiag_idx][J3_col_offdiag_idx] = -Vm[b1-1] * Vm[b2-1] * Ym[b1-1][b2-1] * math.cos((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                
                                    if self.param.BUS[b1][3] == 1 and self.param.BUS[b2][3] == 1:
                                        # off diagonal elements of J4
                                        A[J4_row_offdiag_idx][J4_col_offdiag_idx] = -Vm[b1-1] * Ym[b1-1][b2-1] * math.sin((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                    #   
                    Pk = Vm[b1-1]**2 * Ym[b1-1][b1-1] * math.cos((theta[b1-1][b1-1])) + J33
                    Qk = -Vm[b1-1]**2 * Ym[b1-1][b1-1] * math.sin((theta[b1-1][b1-1])) - J11
                    if self.param.BUS[b1][3] == 3:
                        # swing bus
                        P[b1-1] = Pk
                        Q[b1-1] = Qk
                    if self.param.BUS[b1][3] == 2:
                        Q[b1-1] = Qk 
                        """
                        if Qmax[b1-1] != 0:
                            Qgc = Q[b1-1] + Qd[n-1] - Qsh[n-1]
                            if ii <= 7:                   #Between the 2th & 6th iterations
                                if ii > 2:                #the Mvar of generator buses are
                                    if Qgc < Qmin[n-1]:   #tested. If not within limits Vm(n)
                                        Vm[n-1] += 0.01   #is changed in steps of 0.01 pu to
                                    elif Qgc > Qmax[n-1]: #bring the generator Mvar within
                                        Vm[n-1] -= 0.01   #the specified limits.            
                        """
                    if self.param.BUS[b1][3] != 3:
                        #diagonal elements of J1
                        A[J1_diag_idx][J1_diag_idx] = J11        
                        DC[J1_diag_idx] = P[b1-1] - Pk
                    if self.param.BUS[b1][3] == 1:
                        #diagonal elements of J2
                        A[J2_row_diag_idx][J2_col_diag_idx] = 2 * Vm[b1-1] * Ym[b1-1][b1-1] * math.cos(theta[b1-1][b1-1]) + J22    
                        #diagonal elements of J3
                        A[J3_row_diag_idx][J3_col_diag_idx] = J33
                        #diagonal elements of J4         
                        A[J4_diag_idx][J4_diag_idx] = -2 * Vm[b1-1] * Ym[b1-1][b1-1] * math.sin(theta[b1-1][b1-1]) - J44   
                        DC[J4_diag_idx] = Q[b1-1] - Qk
                #matrix A left division Matrix DC
                DX = np.linalg.solve(A, DC.T)
                for bi in self.config.setBusHnd:
                    del_update_DXidx = int(bi - slackCounted[bi-1]-1)
                    Vm_update_DXidx = int(self.param.nBus + bi - PVcounted[bi-1] - slackCounted[bi-1] - self.param.nSlack-1)
                    if self.param.BUS[bi][3] != 3:
                        delta[bi-1] +=  DX[del_update_DXidx]
                    if self.param.BUS[bi][3] == 1:
                        Vm[bi-1] += DX[Vm_update_DXidx]
                maxerror = max(abs(DC))
                if maxerror < self.param.epsilon:
                    break
                if ii==self.param.iterMax:
                    return {'FLAG':'DIVERGENCE'}
            #finish NR
            # store all voltage magnitude value in all profile
            va.extend(Vm)
            V = [Vm[bi-1]* np.cos(delta[bi-1])+Vm[bi-1]*1j*np.sin(delta[bi-1]) for bi in self.config.busC ]
            Il = dict()
            slt = 0
            vbus,sbus = dict(),dict()
            for i in range(len(Vm)):
                vbus[i+1] = V[i]
                sbus[i+1] = P[i] + Q[i]*1j
            for bi in self.param.busSlack:
                sbus[bi] += self.param.loadProfile[pi][bi]
            for li,bi in self.config.lineC.items():
                Ib1 = (vbus[bi[0]]-vbus[bi[1]])*(-Ybus[bi[0]-1][bi[1]-1])
                Ib2 = (vbus[bi[1]]-vbus[bi[0]])*(-Ybus[bi[0]-1][bi[1]-1])
                #
                if abs(Ib1) >= abs(Ib2):
                    Il[li] = abs(Ib1)
                else: 
                    Il[li] = abs(Ib2)   
                rate = ((Il[li])/self.param.LINE[li][3])*RATEC
                ra.append(rate)
                Snk = V[bi[0]-1]*np.conj(Ib1)
                Skn = V[bi[1]-1]*np.conj(Ib2)
                slt += Snk +Skn  
            for bs in self.param.busSlack:
                if sbus[bs].imag:
                    cosP.append(sbus[bs].real/abs(sbus[bs]))
                else:
                    cosN.append(-sbus[bs].real/abs(sbus[bs]))
            if fo:
                sa1,va1,dia1 = dict(),dict(),dict()# for 1 profile
                va1.update(vbus)
                dia1.update(Il)
                sa1.update(sbus)
                rB,rL,rG =super().__update1profile__(pi,rB,rL,rG,sa1,va1,dia1)
            res['DeltaA'] += slt.real
        res = super().__update_result__(res,va,ra,cosP,cosN)
        if fo:
            super().__export_profiles__(fo,rB,rL,rG,res)
        return res

class PowerSummation(RunMethod):
    def __init__(self,param:Parameter,config:Configuration):
        super().__init__(param=param,config=config)
        self.busGroup = []# cac bus tuong ung o cac slack khac nhau
        for bs1 in self.param.busSlack:
            r1 = self.config.__findBusConnected__(bs1,self.config.busC,self.config.lineC)
            #
            self.busGroup.append(r1)
    #
    def __lineDirection__(self):
        ba = self.param.busSlack[:]
        lset = set()
        for ii in range(20000):
            ba2 = []
            for b1 in ba:
                for l1 in self.config.setLineHnd:
                    if l1 not in lset:
                        if b1==self.config.lineC[l1][1]:
                            d = self.config.lineC[l1][0]
                            self.config.lineC[l1][0] = self.config.lineC[l1][1]
                            self.config.lineC[l1][1] = d
                            lset.add(l1)
                            ba2.append(d)
                        elif b1==self.config.lineC[l1][0]:
                            lset.add(l1)
                            ba2.append(self.config.lineC[l1][1])
            if len(ba2)==0:
                break
            ba= ba2.copy()
    #
    def __ordCompute__(self):
        busC = dict() # connect [LineUp,[LineDown]]
        for h1 in self.config.setBusHnd:
            busC[h1] = [0,set()]
        #
        for h1,l1 in self.config.lineC.items():
            busC[l1[1]][0]= h1     # frombus
            busC[l1[0]][1].add(h1) # tobus
        #
        self.ordc,self.ordv = [],[]
        for bs1 in self.busGroup:
            busC1 = {k:v for k,v in busC.items() if k in bs1}
            #bus already
            balr = {h1:True for h1 in bs1}
            #set order
            sord = set()
            #order compute
            ordc1 = []
            for k,v in busC1.items():
                if len(v[1])==0:
                    if v[0]!=0:
                        ordc1.append(v[0])
                        sord.add(v[0])
                        balr[k]=False
            #
            for ii in range(500):
                for k,v in busC1.items():
                    if balr[k]:
                        if len(v[1]-sord)==0:
                            if k in self.param.setSlack:
                                break
                            #
                            if v[0]!=0:
                                ordc1.append(v[0])
                            sord.add(v[0])
                            balr[k]=False
            ordv1 = [ordc1[-i-1]  for i in range(len(ordc1))]
            self.ordc.append(ordc1)
            self.ordv.append(ordv1)
    #
    def __run1config__(self,fo=''):
        """
        - result (dict): {'FLAG':,'RateMax%', 'Umax[pu]','Umin[pu]','DeltaA','RateMax%'}
        - FLAG (str): 'CONVERGENCE' or 'DIVERGENCE' or 'LOOP' or 'ISLAND'
        - DeltaA: MWH
        """
        # ok run PSM
        self.__lineDirection__()
        #print(self.lineC)
        self.__ordCompute__()
        #print(self.ordc)
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
        # B of Line
        BUSb = {}
        for bri,v in self.param.LINEb.items():
            if bri not in self.config.lineOff:
                bfrom = self.config.lineC[bri][0]
                bto = self.config.lineC[bri][1]
                #
                if bfrom in BUSb.keys():
                    BUSb[bfrom]+=v
                else:
                    BUSb[bfrom]=v
                #
                if bto in BUSb.keys():
                    BUSb[bto]+=v
                else:
                    BUSb[bto]=v

        # Shunt
        for k1,v1 in self.param.BUSbs.items():
            if k1 not in self.config.shuntOff:
                if k1 in BUSb.keys():
                    BUSb[k1]+=v1
                else:
                    BUSb[k1]=v1
        #
        if fo:
            rB,rG,rL = super().__initcsv__(fo)
        #
        va,ra,cosP,cosN = [],[],[1],[-1]
        for pi in self.param.profileID:
            res['DeltaA']-=self.param.loadAll[pi].real
            sa1,va1,dia1 = dict(),dict(),dict()# for 1 profile
            for i1 in range(self.param.nSlack):# with each slack bus
                bs1 = self.param.busSlack[i1]
                ordc1 = self.ordc[i1]
                ordv1 = self.ordv[i1]
                setBusHnd1 = self.busGroup[i1]
                vbus = {h1:complex(self.param.Ubase,0) for h1 in setBusHnd1}
                vbus[bs1] = complex(self.param.genProfile[pi][bs1],0)
                #
                du,di = dict(),dict()
                s0 = 0
                for ii in range(self.param.iterMax+1):
                    sbus = {k:v for k,v in self.param.loadProfile[pi].items() if k in setBusHnd1}
                    # B of Line + Shunt
                    for k1,v1 in BUSb.items():
                        if k1 in setBusHnd1:
                            vv = abs(vbus[k1])
                            sbus[k1] += complex(0, -vv*vv*v1)
                    # cal cong suat nguoc
                    for bri in ordc1:
                        bfrom = self.config.lineC[bri][0]
                        bto = self.config.lineC[bri][1]
                        rx = self.param.LINE[bri][2]
                        #
                        du[bri] = sbus[bto].conjugate()/vbus[bto].conjugate()*rx
                        ib = abs(sbus[bto]/vbus[bto])
                        di[bri] = ib
                        ds1 = ib*ib*rx
                        #
                        if ds1.real>0.2 and ds1.real>sbus[bto].real:# neu ton that lon hon cong suat cua tai
                            return {'FLAG':'DIVERGENCE'}
                        #
                        sbus[bfrom]+=ds1+sbus[bto]
                    # cal dien ap xuoi
                    for bri in ordv1:
                        bfrom = self.config.lineC[bri][0]
                        bto = self.config.lineC[bri][1]
                        vbus[bto]=vbus[bfrom]-du[bri]
                    #
                    if abs(s0-sbus[bs1])<self.param.epsilon:
                        break
                    else:
                        s0 = sbus[bs1]
                    #
                    if ii==self.param.iterMax:
                        return {'FLAG':'DIVERGENCE'}
                # finish
                # loss P
                res['DeltaA']+=sbus[bs1].real
                # Umax[pu]/Umin[pu]
                va.extend( [abs(v) for v in vbus.values()] )
                #
                try:
                    if sbus[bs1].imag>=0:
                        cosP.append(sbus[bs1].real/abs(sbus[bs1]))
                    else:
                        cosN.append(-sbus[bs1].real/abs(sbus[bs1]))
                except:
                    pass
                # RateMax
                for bri in ordc1:
                    ra.append( di[bri]/self.param.LINE[bri][3]*RATEC )
                #
                if fo:
                    va1.update(vbus)
                    dia1.update(di)
                    sa1.update(sbus)
            #
            if fo:
                rB,rL,rG = super().__update1profile__(pi,rB,rL,rG,sa1,va1,dia1)
        #
        res = super().__update_result__(res,va,ra,cosP,cosN)
        #
        if fo:
            super().__export_profiles__
        return res

class PowerFlow:
    def __init__(self,param: Parameter,config: Configuration) -> None:
        self.config = config
        self.param = param
        return
    
    def run1Config_WithObjective(self,option=None,fo=''):
        #
        v1 = self.run1Config(fo)
        if v1['FLAG']!='CONVERGENCE':
            obj = math.inf
        else: #RateMax[%]    Umax[pu]    Umin[pu]    Algo_PF    option_PF
            obj = v1['DeltaA']
            # constraint
            obj+=self.param.setting['RateMax[%]'][1]*max(0, v1['RateMax[%]']-self.param.setting['RateMax[%]'][0])
            obj+=self.param.setting['Umax[pu]'][1]*max(0, v1['Umax[pu]']-self.param.setting['Umax[pu]'][0])
            obj+=self.param.setting['Umin[pu]'][1]*max(0,-v1['Umin[pu]']+self.param.setting['Umin[pu]'][0])
            #cosP ycau cosP>0.9
            obj+=self.param.setting['cosPhiP'][1]*max(0,-v1['cosP']+self.param.setting['cosPhiP'][0])
            #cosN ycau cosN<-0.95
            obj+=self.param.setting['cosPhiN'][1]*max(0,v1['cosN']-self.param.setting['cosPhiN'][0])
        #
        #self.config.lineOff.sort()
        #self.config.shuntOff.sort()
        v1['Objective'] = obj
        v1['LineOff'] = self.config.lineOff
        v1['ShuntOff'] = self.config.shuntOff
        return v1
    
    def run1Config(self,fo=''):
        #check loop island and return if loop or island appears 
        t0 = time.time()
        #
        c1 = self.config.__checkLoopIsland__(self.config.lineOff)
        if c1:
            return {'FLAG':c1}
        """ run PF 1 config """
        if self.param.setting['Algo_PF']=='PSM':
            return PowerSummation(self.param,self.config).__run1config__()  
        elif self.param.setting['Algo_PF']=='N-R':
            return NewtonRaphson(self.param,self.config).__run1config__() 
        elif self.param.setting['Algo_PF'] == 'GS':
            return GaussSeidel(self.param,self.config).__run1config__()
        elif self.param.setting['Algo_PF'] == 'SNR':
            return SparseNewtonRaphson(self.param,self.config).__run1config__()
        return None


if __name__ == "__main__":
    ARGVS.fi = 'Inputs12_2.xlsx'
    ARGVS.fo = 'res\\res1Config.csv'
    param = Parameter(ARGVS.fi)
    lineOff = [3,12,13,14,15,16]
    config = Configuration(param=param,lineOff=lineOff)
    pf = PowerFlow(param=param,config=config)
    res = pf.run1Config_WithObjective(fo=ARGVS.fo)
    print(res)

