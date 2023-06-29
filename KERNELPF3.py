__author__    = "Dr. Pham Quang Phuong"
__copyright__ = "Copyright 2023"
__license__   = "All rights reserved"
__email__     = "phuong.phamquang@hust.edu.vn"
__status__    = "Released"
__version__   = "1.1.5"
"""
about: ....
"""
import os,time,math
import openpyxl,csv
import argparse
import numpy as np 
PARSER_INPUTS = argparse.ArgumentParser(epilog= "")
PARSER_INPUTS.usage = 'Distribution network analysis Tools'
PARSER_INPUTS.add_argument('-fi' , help = '*(str) Input file path' , default = '',type=str,metavar='')
PARSER_INPUTS.add_argument('-fo' , help = '*(str) Output file path', default = '',type=str,metavar='')
ARGVS = PARSER_INPUTS.parse_known_args()[0]
#
RATEC = 100/math.sqrt(3)

#
def toString(v,nRound=5):
    """ convert object/value to String """
    if v is None:
        return 'None'
    t = type(v)
    if t==str:
        if "'" in v:
            return ('"'+v+'"').replace('\n',' ')
        return ("'"+v+"'").replace('\n',' ')
    if t==int:
        return str(v)
    if t==float:
        if v>1.0:
            s1 = str(round(v,nRound))
            return s1[:-2] if s1.endswith('.0') else s1
        elif abs(v)<1e-8:
            return '0'
        s1 ='%.'+str(nRound)+'g'
        return s1 % v
    if t==complex:
        if v.imag>=0:
            return '('+ toString(v.real,nRound)+' +' + toString(v.imag,nRound)+'j)'
        return '('+ toString(v.real,nRound) +' '+ toString(v.imag,nRound)+'j)'
    try:
        return v.toString()
    except:
        pass
    if t in {list,tuple,set}:
        s1=''
        for v1 in v:
            s1+=toString(v1,nRound)+','
        if v:
            s1 = s1[:-1]
        if t==list:
            return '['+s1+']'
        elif t==tuple:
            return '('+s1+')'
        else:
            return '{'+s1+'}'
    if t==dict:
        s1=''
        for k1,v1 in v.items():
            s1+=toString(k1)+':'
            s1+=toString(v1,nRound)+','
        if s1:
            s1 = s1[:-1]
        return '{'+s1+'}'
    return str(v)
#
def __checkLoop__(busHnd,bus,br):
    setBusChecked = set()
    setBrChecked = set()
    for o1 in busHnd:
        if o1 not in setBusChecked:
            setBusChecked.add(o1)
            tb1 = {o1}
            #
            for i in range(20000):
                if i==19999:
                    raise Exception('Error in checkLoop()')
                tb2 = set()
                for b1 in tb1:
                    for l1 in bus[b1]:
                        if l1 not in setBrChecked:
                            setBrChecked.add(l1)
                            for bi in br[l1]:
                                if bi!=b1:
                                    if bi in setBusChecked or bi in tb2:
                                        return bi
                                    tb2.add(bi)
                if len(tb2)==0:
                    break #ok finish no loop for this group
                setBusChecked.update(tb2)
                tb1=tb2.copy()
    return None
#
def __findBusConnected__(bi1,bset,lset):
    ## find all bus connected to Bus b1
    res = {bi1}
    ba = {bi1}
    while True:
        ba2 = set()
        for b1 in ba:
            for li in bset[b1]:
                for bi in lset[li]:
                    if bi not in res:
                        ba2.add(bi)
        if ba2:
            res.update(ba2)
            ba=ba2
        else:
            break
    return res
#
def __getLineISL__(busC0):
    lineISL = set() # line can not be off => ISLAND
    busC = busC0.copy()
    while True:
        n1 = len(lineISL)
        for k,v in busC.items():
            if len(v)==1:
                lineISL.update(v)
        if n1==len(lineISL):
            break
        busc1 = dict()
        for k,v in busC.items():
            if len(v)!=1:
                busc1[k]=v-lineISL
        busC = busc1.copy()
    return lineISL,busC
#
def add2CSV(nameFile,ares,delim):
    """
    append array String to a file CSV
    """
    pathdir = os.path.split(os.path.abspath(nameFile))[0]
    try:
        os.mkdir(pathdir)
    except:
        pass
    #
    if not os.path.isfile(nameFile):
        with open(nameFile, mode='w') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n")
            for a1 in ares:
                ew.writerow(a1)
            f.close()
    else:
        with open(nameFile, mode='a') as f:
            ew = csv.writer(f, delimiter=delim, quotechar='"',lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
            for a1 in ares:
                ew.writerow(a1)
#
class POWERFLOW:
    #assume that bus number are increasing
    def __init__(self,fi):
        self.tcheck = 0
        wbInput = openpyxl.load_workbook(os.path.abspath(fi),data_only=True)
        loadProfile = self.__readInput1Sheet__(wbInput,'LOADPROFILE')
        genProfile = self.__readInput1Sheet__(wbInput,'GENPROFILE')
        busa = self.__readInput1Sheet__(wbInput,'BUS')
        linea = self.__readInput1Sheet__(wbInput,'LINE')
        self.setting = self.__readSetting__(wbInput)
        #
        self.iterMax = int(self.setting['option_PF'][0])
        self.epsilon = self.setting['option_PF'][1]
        self.lineOff = None
        # print(self.setting)
        # BUS[NO] =[ kV,PLOAD[kw],QLOAD[kvar],code ]
        self.BUS = {}
        self.busSlack = []
        self.BUSbs = {} #shunt
        self.Qsh = {}
        for i in range(len(busa['NO'])):
            if busa['FLAG'][i]:
                n1 = busa['NO'][i]
                kv = busa['kV'][i]
                p1 = busa['PLOAD[kw]'][i]/1000
                q1 = busa['QLOAD[kvar]'][i]/1000
                qsh = busa['Qshunt[kvar]'][i]/1000
                if abs(qsh)>1e-6:
                    self.BUSbs[n1] = qsh
                #
                c1 = busa['CODE'][i]
                if c1==None:
                    c1=1
                if c1 in {2,3}:
                    self.busSlack.append(n1)
                #
                self.BUS[n1] = [kv,p1,q1,c1]
        self.nSlack = len(self.busSlack)
        self.setSlack = set(self.busSlack)
        #
        self.Ubase = self.BUS[self.busSlack[0]][0]
        self.Ubase2 = self.Ubase*self.Ubase
        # update B shunt at bus
        for k1,v1 in self.BUSbs.items():
            self.BUSbs[k1] = v1/self.Ubase2# q =u*u*b
        #
        self.profileID = [int(i) for i in loadProfile['time\\NOBUS']]
        # LOAD PROFILE convert to MVA
        self.loadProfile = dict()
        self.loadAll = dict()
        for ii in range(len(self.profileID)):
            k = loadProfile['time\\NOBUS'][ii]
            v1 = dict()
            self.loadAll[k] = 0
            for k1 in loadProfile.keys():
                if k1!='time\\NOBUS':
                    n1 = int(k1)
                    v1[n1] = loadProfile[k1][ii] * complex(self.BUS[n1][1],self.BUS[n1][2])
                    self.loadAll[k]+=v1[n1]
            self.loadProfile[k] = v1
        # GENPROFILE convert to kV
        self.genProfile = dict()
        for ii in range(len(self.profileID)):
            k = genProfile['time\\NOBUS'][ii]
            v1 = dict()
            for k1 in genProfile.keys():
                if k1!='time\\NOBUS':
                    n1 = int(k1)
                    v1[n1] = genProfile[k1][ii] * self.Ubase
            self.genProfile[k]=v1
        # LINE[NO] = [FROMBUS,TOBUS,RX(Ohm),B/2(Siemens),RATE ]
        self.LINE = {}
        self.LINEb = {} # b of Line
        for i in range(len(linea['NO'])):
            if linea['FLAG'][i]:
                n1= linea['NO'][i]
                fr = linea['FROMBUS'][i]
                to = linea['TOBUS'][i]
                r = linea['R(Ohm)'][i]
                x = linea['X(Ohm)'][i]
                r1 = linea['RATEA[A]'][i]/1000 #kA
                self.LINE[n1] = [fr,to,complex(r,x),r1]
                #
                if linea['B(microSiemens)'][i]>1e-2:
                    self.LINEb[n1] = linea['B(microSiemens)'][i]*1e-6/2
        self.setLineHndAll = set(self.LINE.keys())
        #
        self.setBusHnd = set(self.BUS.keys())
        self.lstBusHnd = busa['NO']
        self.lstLineHnd= linea['NO']
        # list cac line co the dong mo
        self.lineFLAG3 = []
        for i in range(len(linea['NO'])):
            if linea['FLAG3'][i]:
                self.lineFLAG3.append(linea['NO'][i])
        #
        self.shuntFLAG3 = []
        if 'FLAG3' in busa.keys():
            for i in range(len(busa['NO'])):
                if busa['FLAG3'][i]:
                    self.shuntFLAG3.append(busa['NO'][i])
        #
        self.nVar = len(self.lineFLAG3) + len(self.shuntFLAG3)
        self.nL = len(self.lineFLAG3)
        #
        self.BUSC = dict() #connect of BUS
        for b1 in self.setBusHnd:
            self.BUSC[b1] = set()
        #
        for k,v in self.LINE.items():
            self.BUSC[v[0]].add(k)
            self.BUSC[v[1]].add(k)
        #
        self.t0 = time.time()
        #print(self.BUSC)
        self.lineSureISL,busc1 = __getLineISL__(self.BUSC)
        self.bus0ISL = set(busc1.keys())
        self.busISL = self.setBusHnd - self.bus0ISL
        #print(self.lineISL) # line ko the off, off=>island
        #assume that only take circumstances where there is no island bus to calculate power flow
        self.nBus = len(self.setBusHnd)
        #print('busISL',self.busISL)   # bus con lai sau khi da bo line island, dung de check loop
    #
    def getLineFlag3(self):
        """ cac Branch co the dong mo """
        return self.lineFLAG3
    #
    def getShuntFlag3(self):
        """ cac Shunt co the dong mo """
        return self.shuntFLAG3
    #
    def getLineOff(self,lineFlag): # 0: inservice, 1 off service
        lineOff = []
        for i in range(len(self.lineFLAG3)):
            if lineFlag[i]:
                lineOff.append(self.lineFLAG3[i])
        return lineOff
    #
    def getShuntOff(self,shuntFlag): # 0: inservice, 1 off service
        shuntOff = []
        for i in range(len(self.shuntFLAG3)):
            if shuntFlag[i]:
                shuntOff.append(self.shuntFLAG3[i])
        return shuntOff
    #
    def getVarFlag(self,lineOff,shuntOff):
        varFlag = [0]*self.nVar
        for i in range(self.nL):
            if self.lineFLAG3[i] in lineOff:
                varFlag[i]=1
        for i in range(self.nVar-self.nL):
            if self.shuntFLAG3[i] in shuntOff:
                varFlag[self.nL+i]=1
        return varFlag
    #
    def run1Config_WithObjective(self,lineOff=[],shuntOff=[],varFlag=None,option=None,fo=''):
        if varFlag is not None:
            if len(varFlag)!=self.nVar:
                raise Exception('Error size of varFlag')
            lineOff = self.getLineOff(varFlag[:self.nL])
            shuntOff = self.getShuntOff(varFlag[self.nL:])
        #
        v1 = self.run1Config(set(lineOff),set(shuntOff),fo)
        if v1['FLAG']!='CONVERGENCE':
            obj = math.inf
        else: #RateMax[%]    Umax[pu]    Umin[pu]    Algo_PF    option_PF
            obj = v1['DeltaA']
            # constraint
            obj+=self.setting['RateMax[%]'][1]*max(0, v1['RateMax[%]']-self.setting['RateMax[%]'][0])
            obj+=self.setting['Umax[pu]'][1]*max(0, v1['Umax[pu]']-self.setting['Umax[pu]'][0])
            obj+=self.setting['Umin[pu]'][1]*max(0,-v1['Umin[pu]']+self.setting['Umin[pu]'][0])
            #cosP ycau cosP>0.9
            obj+=self.setting['cosPhiP'][1]*max(0,-v1['cosP']+self.setting['cosPhiP'][0])
            #cosN ycau cosN<-0.95
            obj+=self.setting['cosPhiN'][1]*max(0,v1['cosN']-self.setting['cosPhiN'][0])
        #
        lineOff.sort()
        shuntOff.sort()
        v1['Objective'] = obj
        v1['LineOff'] = lineOff
        v1['ShuntOff'] = shuntOff
        return v1
    #
    def run1Config(self,lineOff=set(),shuntOff=set(),fo=''):
        #check loop island and return if loop or island appears 
        t0 = time.time()
        #
        c1 = self.__checkLoopIsland__(lineOff)
        self.tcheck+=time.time()-t0
        if c1:
            return {'FLAG':c1}
        """ run PF 1 config """
        if self.setting['Algo_PF']=='PSM':
            return self.__run1ConfigPSM__(lineOff,shuntOff,fo)  
        elif self.setting['Algo_PF']=='N-R':
            return self.__run1configNR__(lineOff,shuntOff,fo)
            
        elif self.setting['Algo_PF'] == 'GS':
            return self.__run1configGS__(lineOff,shuntOff,fo)
        elif self.setting['Algo_PF'] == 'SNR':
            return self.__run1configSNR__(lineOff,shuntOff,fo)
        return None
        
    #
    def __run1ConfigPSM__(self,lineOff,shuntOff,fo=''):
        """
        - result (dict): {'FLAG':,'RateMax%', 'Umax[pu]','Umin[pu]','DeltaA','RateMax%'}
        - FLAG (str): 'CONVERGENCE' or 'DIVERGENCE' or 'LOOP' or 'ISLAND'
        - DeltaA: MWH
        """
        # ok run PSM
        self.__lineDirection__()
        #print(self.lineC)
        self.__ordCompute__()
        #print(self.ordc)
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
        # B of Line
        BUSb = {}
        for bri,v in self.LINEb.items():
            if bri not in lineOff:
                bfrom = self.lineC[bri][0]
                bto = self.lineC[bri][1]
                #
                if bfrom in BUSb.keys():
                    BUSb[bfrom]+=v
                else:
                    BUSb[bfrom]=v
                #
                if bto in BUSb.keys():
                    BUSb[bto]+=v
                else:
                    BUSb[bto]=v

        # Shunt
        for k1,v1 in self.BUSbs.items():
            if k1 not in shuntOff:
                if k1 in BUSb.keys():
                    BUSb[k1]+=v1
                else:
                    BUSb[k1]=v1
        #
        if fo:
            add2CSV(fo,[[],[time.ctime()],['PF 1Profile','lineOff',str(list(lineOff)),'shuntOff',str(list(shuntOff))]],',')
            #
            rB = [[],['BUS/Profile']]
            rB[1].extend([bi for bi in self.lstBusHnd])
            #
            rL = [[],['LINE/Profile']]
            rL[1].extend([bi for bi in self.lstLineHnd])
            #
            rG = [[],['GEN/Profile']]
            for bi in self.busSlack:
                 rG[1].append(str(bi)+'_P')
                 rG[1].append(str(bi)+'_Q')
                 rG[1].append(str(bi)+'_cosPhi')
        #
        va,ra,cosP,cosN = [],[],[1],[-1]
        for pi in self.profileID:
            res['DeltaA']-=self.loadAll[pi].real
            sa1,va1,dia1 = dict(),dict(),dict()# for 1 profile
            for i1 in range(self.nSlack):# with each slack bus
                bs1 = self.busSlack[i1]
                ordc1 = self.ordc[i1]
                ordv1 = self.ordv[i1]
                setBusHnd1 = self.busGroup[i1]
                vbus = {h1:complex(self.Ubase,0) for h1 in setBusHnd1}
                vbus[bs1] = complex(self.genProfile[pi][bs1],0)
                #
                du,di = dict(),dict()
                s0 = 0
                for ii in range(self.iterMax+1):
                    sbus = {k:v for k,v in self.loadProfile[pi].items() if k in setBusHnd1}
                    # B of Line + Shunt
                    for k1,v1 in BUSb.items():
                        if k1 in setBusHnd1:
                            vv = abs(vbus[k1])
                            sbus[k1] += complex(0, -vv*vv*v1)
                    # cal cong suat nguoc
                    for bri in ordc1:
                        bfrom = self.lineC[bri][0]
                        bto = self.lineC[bri][1]
                        rx = self.LINE[bri][2]
                        #
                        du[bri] = sbus[bto].conjugate()/vbus[bto].conjugate()*rx
                        ib = abs(sbus[bto]/vbus[bto])
                        di[bri] = ib
                        ds1 = ib*ib*rx
                        #
                        if ds1.real>0.2 and ds1.real>sbus[bto].real:# neu ton that lon hon cong suat cua tai
                            return {'FLAG':'DIVERGENCE'}
                        #
                        sbus[bfrom]+=ds1+sbus[bto]
                    # cal dien ap xuoi
                    for bri in ordv1:
                        bfrom = self.lineC[bri][0]
                        bto = self.lineC[bri][1]
                        vbus[bto]=vbus[bfrom]-du[bri]
                    #
                    if abs(s0-sbus[bs1])<self.epsilon:
                        break
                    else:
                        s0 = sbus[bs1]
                    #
                    if ii==self.iterMax:
                        return {'FLAG':'DIVERGENCE'}
                # finish
                # loss P
                res['DeltaA']+=sbus[bs1].real
                # Umax[pu]/Umin[pu]
                va.extend( [abs(v) for v in vbus.values()] )
                #
                try:
                    if sbus[bs1].imag>=0:
                        cosP.append(sbus[bs1].real/abs(sbus[bs1]))
                    else:
                        cosN.append(-sbus[bs1].real/abs(sbus[bs1]))
                except:
                    pass
                # RateMax
                for bri in ordc1:
                    ra.append( di[bri]/self.LINE[bri][3]*RATEC )
                #
                if fo:
                    va1.update(vbus)
                    dia1.update(di)
                    sa1.update(sbus)
            #
            if fo:
                rb1 = [pi]
                rl1 = [pi]
                rg1 = [pi]
                for bi1 in self.lstBusHnd:
                    rb1.append(toString(abs(va1[bi1])/self.Ubase))
                #
                for bri in self.lstLineHnd:
                    try:
                        r1 = dia1[bri]/self.LINE[bri][3]*RATEC
                        rl1.append( toString(r1,2) )
                    except:
                        rl1.append('0')
                #
                for bs1 in self.busSlack:
                    rg1.append(toString(sa1[bs1].real))
                    rg1.append(toString(sa1[bs1].imag))
                    if sa1[bs1].imag>=0:
                        rg1.append(toString(sa1[bs1].real/abs(sa1[bs1]),3))
                    else:
                        rg1.append(toString(-sa1[bs1].real/abs(sa1[bs1]),3))
                #
                rB.append(rb1)
                rL.append(rl1)
                rG.append(rg1)
        #
        va.sort()
        res['Umax[pu]'] = va[-1]/self.Ubase
        res['Umin[pu]'] = va[0]/self.Ubase
        res['RateMax[%]'] = max(ra)
        res['cosP'] = min(cosP)
        res['cosN'] = max(cosN)
        #
        if fo:
            rB.append(['','Umax[pu]',toString(res['Umax[pu]']),'Umin[pu]',toString(res['Umin[pu]']) ])
            add2CSV(fo,rB,',')
            #
            rL.append(['','RateMax[%]',toString(res['RateMax[%]'],2)])
            add2CSV(fo,rL,',')
            #
            rG.append(['','cosPmin',toString(res['cosP'],3),'cosNMax',toString(res['cosN'],3)])
            add2CSV(fo,rG,',')
        #
        return res
    #
    def __ordCompute__(self):
        busC = dict() # connect [LineUp,[LineDown]]
        for h1 in self.setBusHnd:
            busC[h1] = [0,set()]
        #
        for h1,l1 in self.lineC.items():
            busC[l1[1]][0]= h1     # frombus
            busC[l1[0]][1].add(h1) # tobus
        #
        self.ordc,self.ordv = [],[]
        for bs1 in self.busGroup:
            busC1 = {k:v for k,v in busC.items() if k in bs1}
            #bus already
            balr = {h1:True for h1 in bs1}
            #set order
            sord = set()
            #order compute
            ordc1 = []
            for k,v in busC1.items():
                if len(v[1])==0:
                    if v[0]!=0:
                        ordc1.append(v[0])
                        sord.add(v[0])
                        balr[k]=False
            #
            for ii in range(500):
                for k,v in busC1.items():
                    if balr[k]:
                        if len(v[1]-sord)==0:
                            if k in self.setSlack:
                                break
                            #
                            if v[0]!=0:
                                ordc1.append(v[0])
                            sord.add(v[0])
                            balr[k]=False
            ordv1 = [ordc1[-i-1]  for i in range(len(ordc1))]
            self.ordc.append(ordc1)
            self.ordv.append(ordv1)
    #
    def __lineDirection__(self):
        ba = self.busSlack[:]
        lset = set()
        for ii in range(20000):
            ba2 = []
            for b1 in ba:
                for l1 in self.setLineHnd:
                    if l1 not in lset:
                        if b1==self.lineC[l1][1]:
                            d = self.lineC[l1][0]
                            self.lineC[l1][0] = self.lineC[l1][1]
                            self.lineC[l1][1] = d
                            lset.add(l1)
                            ba2.append(d)
                        elif b1==self.lineC[l1][0]:
                            lset.add(l1)
                            ba2.append(self.lineC[l1][1])
            if len(ba2)==0:
                break
            ba= ba2.copy()
    #
    def __checkLoopIsland__(self,lineOff):
        # check island/loop multi slack ----------------------------------------
        if lineOff.intersection(self.lineSureISL):
            return 'ISLAND'
        #
        self.setLineHnd = self.setLineHndAll - lineOff
        #
        self.lineC = {k:self.LINE[k][:2] for k in self.setLineHnd}
        self.busC = {b1:set() for b1 in self.setBusHnd}
        for k,v in self.lineC.items():
            self.busC[v[0]].add(k)
            self.busC[v[1]].add(k)
        #
        r11 = self.setBusHnd.copy()
        self.busGroup = []# cac bus tuong ung o cac slack khac nhau
        for bs1 in self.busSlack:
            r1 = __findBusConnected__(bs1,self.busC,self.lineC)
            if len(r1.intersection(self.setSlack))>1:
                return 'LOOP MULTI SLACK'
            #
            self.busGroup.append(r1)
            r11.difference_update(r1)
        #
        if r11:
            return 'ISLAND'
        #
        # LOOP
        if __checkLoop__(self.bus0ISL,self.busC,self.lineC):
            return 'LOOP'
        #
        return ''
    #
    def __readInput1Sheet__(self,wbInput,sh1):
        ws = wbInput[sh1]
        res = {}
        # dem so dong data
        for i in range(2,20000):
            if ws.cell(i,1).value==None:
                k=i
                break
        #
        for i in range(1,20000):
            v1 = ws.cell(2,i).value
            if v1==None:
                return res
            va = []
            #
            for i1 in range(3,k):
                va.append(ws.cell(i1,i).value)
            res[str(v1)]=va
        return res
    #
    def __readSetting__(self,wbInput):
        ws = wbInput['SETTING']
        k = 0
        res = {}
        while True:
            k+=1
            s1= ws.cell(k,1).value
            if type(s1)==str and s1.replace(' ','')=='##BRANCHING':
                for j in range(1,100):
                    s2 = str(ws.cell(k+1,j).value).strip()
                    if s2=='None':
                        break
                    sa = str(ws.cell(k+2,j).value).split(',')
                    if len(sa)==1:
                        try:
                            res[s2] = float(sa[0])
                        except:
                            res[s2] = sa[0]
                    else:
                        res[s2] = [float(si) for si in sa]
                break
        return res
    
    def __calculateYbus__(self,lineOff,shuntOff):
        self.setLineHnd = self.setLineHndAll - lineOff
        self.setLinebHnd = self.LINEb.keys() - lineOff
        #initialize Ybus
        Ybus = []
        for _ in range(self.nBus):
            row = [0] * self.nBus
            Ybus.append(row)
        #initialize branch admittance
        y = {k:(1+0j) for k in self.setLineHnd}
        #formation of the off diagonal elements
        for k,v in self.lineC.items():
                y[k] = y[k]/self.LINE[k][2]
                Ybus[v[0]-1][v[1]-1] = -y[k]
                Ybus[v[1]-1][v[0]-1] = Ybus[v[0]-1][v[1]-1]
        #calculate yline with lineb: 
        for lbi in self.setLinebHnd:
            y[lbi] += self.LINEb[lbi]*1j
        #formation of the diagonal elements
        for bi in self.busC.keys():
            for li in self.busC[bi]:
                Ybus[bi-1][bi-1] +=  y[li]
        # Shunt 
        for k1,v1 in self.BUSbs.items():
            if k1 not in shuntOff:
                Ybus[k1-1][k1-1] += v1*1j
        return Ybus

    def __calculate_sparse_Ybus__(self,lineOff,shuntOff):    
        self.setLineHnd = self.setLineHndAll - lineOff
        self.setLinebHnd = self.LINEb.keys() - lineOff
        sparse_ybus = dict()
        #initialize branch admittance
        y = {k:(1+0j) for k in self.setLineHnd}
        #formation of the off diagonal elements
        for k,v in self.lineC.items():
            y[k] = y[k]/self.LINE[k][2]
            sparse_ybus[frozenset({v[0],v[1]})] = -y[k]   
        #calculate yline with lineb: 
        for lbi in self.setLinebHnd:
            y[lbi] += self.LINEb[lbi]*1j
        # Shunt 
        for k1,v1 in self.BUSbs.items():
            if k1 not in shuntOff:
                sparse_ybus[k1] = v1*1j
        #formation of the diagonal elements
        for bi in self.busC.keys():
            for li in self.busC[bi]:
                if bi not in sparse_ybus:
                    sparse_ybus[bi] =  y[li]
                else:
                    sparse_ybus[bi] +=  y[li]
        # case if bus slack does not connect to any bus
        for bsi in self.busSlack:
            if not self.busC[bsi]:
                sparse_ybus[bsi] = 0
        return sparse_ybus

    def __run1configGS__(self,lineOff,shuntOff,fo=''):
        # ready to run Gauss Seidel
        Ybus = self.__calculate_sparse_Ybus__(lineOff,shuntOff)
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
         #
        if fo:
            add2CSV(fo,[[],[time.ctime()],['PF 1Profile','lineOff',str(list(lineOff)),'shuntOff',str(list(shuntOff))]],',')
            #
            rB = [[],['BUS/Profile']]
            rB[1].extend([bi for bi in self.lstBusHnd])
            #
            rL = [[],['LINE/Profile']]
            rL[1].extend([bi for bi in self.lstLineHnd])
            #
            rG = [[],['GEN/Profile']]
            for bi in self.busSlack:
                 rG[1].append(str(bi)+'_P')
                 rG[1].append(str(bi)+'_Q')
                 rG[1].append(str(bi)+'_cosPhi')
        va,ra,cosP,cosN = [],[],[1],[-1]
        accel = self.setting['accel']

        for pi in self.profileID:
            P = {bi:(-self.loadProfile[pi][bi]).real for bi in self.setBusHnd}    #+self.Pgen
            Q = {bi:(-self.loadProfile[pi][bi]).imag for bi in self.setBusHnd}    #+self.Qgen
            DP = {bi:0 for bi in self.setBusHnd}
            DQ = {bi:0 for bi in self.setBusHnd}
            sbus = {bi:(P[bi] + Q[bi]*1j) for bi in self.setBusHnd}
            vbus = {bi:complex(self.Ubase,0) for bi in self.setBusHnd}
            # initialize voltage correction
            Vc = {bi:complex(0,0) for bi in self.setBusHnd}
            # update Vm of slack buses
            for bs in self.setSlack:
                vbus[bs] = complex(self.genProfile[pi][bs],0) 
            Vm = {bi:abs(vbus[bi])for bi in self.setBusHnd}
            sa1,dia1,va1 = dict(),dict(),dict() # for 1 profile
            for ii in range(self.iterMax+1):
                for b1 in self.setBusHnd:
                    YV = 0 + 1j * 0
                    for li in self.busC[b1]:
                        for b2 in self.lineC[li]:
                            if b1 == b2:
                                pass
                            else:
                                line = frozenset({b1,b2})
                                YV += Ybus[line] * vbus[b2]
                    Sc = np.conj(vbus[b1]) * (Ybus[b1] * vbus[b1] + YV)
                    Sc = np.conj(Sc)
                    DP[b1] = P[b1] - np.real(Sc)
                    DQ[b1] = Q[b1] - np.imag(Sc)
                    if self.BUS[b1][3] == 3:
                        sbus[b1] = Sc
                        P[b1] = np.real(Sc)
                        Q[b1] = np.imag(Sc)
                        DP[b1] = 0
                        DQ[b1] = 0
                        Vc[b1] = vbus[b1]
                    elif self.BUS[b1][3] == 2:
                        Q[b1] = np.imag(Sc)
                        sbus[b1] = P[b1] + 1j * Q[b1]
                        """
                        if Qmax[n] != 0:
                            Qgc = Q[n] * basemva + Qd[n] - Qsh[n]
                            if abs(DQ[n]) <= 0.005 and iter >= 10:
                                if DV[n] <= 0.045:
                                    if Qgc < Qmin[n]:
                                        Vm[n] += 0.005
                                        DV[n] += 0.005
                                    elif Qgc > Qmax[n]:
                                        Vm[n] -= 0.005
                                        DV[n] += 0.005
                                else:
                                    pass
                            else:
                                pass
                        else:
                            pass
                        """
                    if self.BUS[b1][3] != 3:
                        Vc[b1] = (np.conj(sbus[b1]) / np.conj(vbus[b1]) - YV) / Ybus[b1]
                    else:
                        pass
                    if self.BUS[b1][3] == 1:
                        vbus[b1] = vbus[b1] + accel * (Vc[b1] - vbus[b1])
                    elif self.BUS[b1][3] == 2:
                        VcI = np.imag(Vc[b1])
                        VcR = np.sqrt(Vm[b1] ** 2 - VcI ** 2)
                        Vc[b1] = VcR + 1j * VcI
                        vbus[b1] = vbus[b1] + accel * (Vc[b1] - vbus[b1])
                maxerror = abs(DP[b1])
                for bi in self.setBusHnd:
                    if abs(DP[bi]) > maxerror:
                        maxerror = abs(DP[bi])
                    if abs(DQ[bi]) > maxerror:
                        maxerror = abs(DQ[bi])
                if maxerror < self.epsilon:
                    break
                if ii==self.iterMax:
                    return {'FLAG':'DIVERGENCE'}
            # finish GS
            # store all voltage magnitude value in all profile
            Vm = {bi:abs(vbus[bi]) for bi in self.setBusHnd}
            va.extend(Vm.values())
            for bi in self.busSlack:
                sbus[bi] += self.loadProfile[pi][bi]
            """
            for bi in self.busPV:
                # update Q load in PV bus
                sbus[bi] += self.loadProfile[pi][bi].imag * 1j 
            """
            slt = 0
            Il = dict()
            for li,bi in self.lineC.items():
                line = frozenset({bi[0],bi[1]})
                Ib1 = (vbus[bi[0]]-vbus[bi[1]])*(-Ybus[line])
                Ib2 = (vbus[bi[1]]-vbus[bi[0]])*(-Ybus[line])
                #
                if abs(Ib1) >= abs(Ib2):
                    Il[li] = abs(Ib1)
                else: 
                    Il[li] = abs(Ib2)   

                rate = ((Il[li])/self.LINE[li][3])*RATEC
                ra.append(rate)
                
                Snk = vbus[bi[0]]*np.conj(Ib1)
                Skn = vbus[bi[1]]*np.conj(Ib2)
                slt += Snk +Skn
            if fo:
                va1.update(vbus)
                dia1.update(Il)
                sa1.update(sbus)
            for bs in self.busSlack:
                if sbus[bs].imag:
                    cosP.append(sbus[bs].real/abs(sbus[bs]))
                else:
                    cosN.append(-sbus[bs].real/abs(sbus[bs]))
            if fo:
                rb1 = [pi]
                rl1 = [pi]
                rg1 = [pi]
                for bi1 in self.lstBusHnd:
                    rb1.append(toString(abs(va1[bi1])/self.Ubase))
                #
                for bri in self.lstLineHnd:
                    try:
                        r1 = abs(Il[bri])/self.LINE[bri][3]*RATEC
                        rl1.append( toString(r1,2) )
                    except:
                        rl1.append('0')
                #
                for bs1 in self.busSlack:
                    rg1.append(toString(sa1[bs1].real))
                    rg1.append(toString(sa1[bs1].imag))
                    if sa1[bs1].imag>=0:
                        rg1.append(toString(sa1[bs1].real/abs(sa1[bs1]),3))
                    else:
                        rg1.append(toString(-sa1[bs1].real/abs(sa1[bs1]),3))
                #
                rB.append(rb1)
                rL.append(rl1)
                rG.append(rg1)
            res['DeltaA'] += slt.real
        res['Umax[pu]'] = max(va)/self.Ubase
        res['Umin[pu]'] = min(va)/self.Ubase
        res['RateMax[%]'] = max(ra)
        res['cosP'] = min(cosP)
        res['cosN'] = max(cosN)
        
        #
        if fo:
            rB.append(['','Umax[pu]',toString(res['Umax[pu]']),'Umin[pu]',toString(res['Umin[pu]']) ])
            add2CSV(fo,rB,',')
            #
            rL.append(['','RateMax[%]',toString(res['RateMax[%]'],2)])
            add2CSV(fo,rL,',')
            #
            rG.append(['','cosPmin',toString(res['cosP'],3),'cosNMax',toString(res['cosN'],3)])
            add2CSV(fo,rG,',')
        #
        return res
             
    def __update1profilevalue__(self,pi=None,fo='',res=[],va1=None,dia1=None,sa1=None,vbus=None,
                                Il=None,sbus=dict(),cosP=None,cosN=None,rB=None,rL=None,rG=None,slt=complex(0,0)):
        if fo:
            va1.update(vbus)
            dia1.update(Il)
            sa1.update(sbus)
        for bs in self.busSlack:
            if sbus[bs].imag:
                cosP.append(sbus[bs].real/abs(sbus[bs]))
            else:
                cosN.append(-sbus[bs].real/abs(sbus[bs]))
        if fo:
            rb1 = [pi]
            rl1 = [pi]
            rg1 = [pi]
            for bi1 in self.lstBusHnd:
                rb1.append(toString(abs(va1[bi1])/self.Ubase))
            #
            for bri in self.lstLineHnd:
                try:
                    r1 = abs(Il[bri])/self.LINE[bri][3]*RATEC
                    rl1.append( toString(r1,2) )
                except:
                    rl1.append('0')
            #
            for bs1 in self.busSlack:
                rg1.append(toString(sa1[bs1].real))
                rg1.append(toString(sa1[bs1].imag))
                if sa1[bs1].imag>=0:
                    rg1.append(toString(sa1[bs1].real/abs(sa1[bs1]),3))
                else:
                    rg1.append(toString(-sa1[bs1].real/abs(sa1[bs1]),3))
            #
            rB.append(rb1)
            rL.append(rl1)
            rG.append(rg1)
        res['DeltaA'] += slt.real
        if fo:
            return res,cosP,cosN,rB,rL,rG
        else:
            return res,cosP,cosN     

    def __run1configSNR__(self,lineOff,shuntOff,fo=''):
        # ready to run Newton raphson
        start = time.time()
        Ybus = self.__calculate_sparse_Ybus__(lineOff,shuntOff)
        end = time.time()
        print(end - start)
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
        #
        if fo:
            add2CSV(fo,[[],[time.ctime()],['PF 1Profile','lineOff',str(list(lineOff)),'shuntOff',str(list(shuntOff))]],',')
            #
            rB = [[],['BUS/Profile']]
            rB[1].extend([bi for bi in self.lstBusHnd])
            #
            rL = [[],['LINE/Profile']]
            rL[1].extend([bi for bi in self.lstLineHnd])
            #
            rG = [[],['GEN/Profile']]
            for bi in self.busSlack:
                 rG[1].append(str(bi)+'_P')
                 rG[1].append(str(bi)+'_Q')
                 rG[1].append(str(bi)+'_cosPhi')
        Ym, theta = dict(),dict()
        for k,v in Ybus.items():
            Ym[k] = abs(v)
            theta[k] = np.angle(v,deg=False)
        #
        countSlack = 0
        countPV = 0
        slackCounted = []
        PVcounted = []
        for bi in self.setBusHnd:
            if self.BUS[bi][3] == 3:
                countSlack+=1 
            elif self.BUS[bi][3] == 2:
                countPV += 1 
            slackCounted.append(countSlack)
            PVcounted.append(countPV)
        no_jacobi_equation = 2 * self.nBus - countPV - 2 * self.nSlack         
        #
        DC = np.zeros(no_jacobi_equation) 
        #
        va,ra,cosP,cosN = [],[],[1],[-1]
        for pi in self.profileID:

            # initialize voltage magnitude of all buses
            # iterate in setBUShnd in case some buses are off 
            Vm = {bi:float(self.Ubase) for bi in self.setBusHnd}
            # initialize voltage angle 
            delta = {bi:0 for bi in self.setBusHnd}
            P = {bi:(-self.loadProfile[pi][bi]).real for bi in self.setBusHnd}    #+self.Pgen
            Q = {bi:(-self.loadProfile[pi][bi]).imag for bi in self.setBusHnd}    #+self.Qgen
            # update Vm of slack buses
            for bs in self.setSlack:
                Vm[bs] = self.genProfile[pi][bs]
            sa1,dia1,va1 = dict(),dict(),dict() # for 1 profile
            for ii in range(self.iterMax+1):
                # Initialize Jacobian Matrix
                A = np.zeros((no_jacobi_equation,no_jacobi_equation))
                for b1 in self.setBusHnd:

                    J1_row_offdiag_idx = J2_row_offdiag_idx = int((b1 - slackCounted[b1-1])-1)
                    J1_diag_idx = J2_row_diag_idx = J3_col_diag_idx = J1_row_offdiag_idx 

                    J3_row_offdiag_idx = J4_row_offdiag_idx = int((self.nBus+b1-slackCounted[b1-1]-PVcounted[b1-1]-self.nSlack)-1)
                    J4_diag_idx = J2_col_diag_idx = J3_row_diag_idx = J3_row_offdiag_idx 

                    J11 = 0
                    J22 = 0
                    J33 = 0
                    J44 = 0
                    #
                    for li in self.busC[b1]:
                        for b2 in self.lineC[li]:
                            if b1 == b2:
                                pass
                            else:
                                line = frozenset({b1,b2})
                                # diagonal elements of J1
                                J11 += Vm[b1] * Vm[b2] * Ym[line] * math.sin((theta[line] - delta[b1] + delta[b2]).real)
                                # diagonal elements of J3          
                                J33 += Vm[b1] * Vm[b2] * Ym[line] * math.cos((theta[line] - delta[b1] + delta[b2]).real)
                                if self.BUS[b1][3] != 3:
                                    #slackbus doesn't exist in J2 and J4
                                    #diagonal elements of J2
                                    J22 += Vm[b2] * Ym[line] * math.cos((theta[line] - delta[b1] + delta[b2]).real)
                                    # diagonal elements of J4               
                                    J44 += Vm[b2] * Ym[line] * math.sin((theta[line] - delta[b1] + delta[b2]).real)
                                if self.BUS[b1][3] != 3 and self.BUS[b2][3] != 3:
                                    J1_col_offdiag_idx = J3_col_offdiag_idx = int(b2 - slackCounted[b2-1] - 1)
                                    J2_col_offdiag_idx = J4_col_offdiag_idx = int(self.nBus + b2 - PVcounted[b2-1] - slackCounted[b2-1] - self.nSlack - 1)
                                    # off diagonal elements of J1
                                    A[J1_row_offdiag_idx][J1_col_offdiag_idx] = -Vm[b1] * Vm[b2] * Ym[line] * math.sin((theta[line] - delta[b1] + delta[b2]).real)
                                
                                    if self.BUS[b2][3] == 1:
                                        # off diagonal elements of J2
                                        A[J2_row_offdiag_idx][J2_col_offdiag_idx] = Vm[b1] * Ym[line] * math.cos((theta[line] - delta[b1] + delta[b2]).real)
                                
                                    if self.BUS[b1][3] == 1:
                                        # off diagonal elements of J3
                                        A[J3_row_offdiag_idx][J3_col_offdiag_idx] = -Vm[b1] * Vm[b2] * Ym[line] * math.cos((theta[line] - delta[b1] + delta[b2]).real)
                                
                                    if self.BUS[b1][3] == 1 and self.BUS[b2][3] == 1:
                                        # off diagonal elements of J4
                                        A[J4_row_offdiag_idx][J4_col_offdiag_idx] = -Vm[b1] * Ym[line] * math.sin((theta[line] - delta[b1] + delta[b2]).real)
                    #   
                    Pk = Vm[b1]**2 * Ym[b1] * math.cos((theta[b1])) + J33
                    Qk = -Vm[b1]**2 * Ym[b1] * math.sin((theta[b1])) - J11
                    if self.BUS[b1][3] == 3:
                        # swing bus
                        P[b1] = Pk
                        Q[b1] = Qk
                    if self.BUS[b1][3] == 2:
                        Q[b1] = Qk
                        """
                        if Qmax[b1-1] != 0:
                            Qgc = Q[b1-1] + Qd[n-1] - Qsh[n-1]
                            if ii <= 7:                   #Between the 2th & 6th iterations
                                if ii > 2:                #the Mvar of generator buses are
                                    if Qgc < Qmin[n-1]:   #tested. If not within limits Vm(n)
                                        Vm[n-1] += 0.01   #is changed in steps of 0.01 pu to
                                    elif Qgc > Qmax[n-1]: #bring the generator Mvar within
                                        Vm[n-1] -= 0.01   #the specified limits.            
                        """
                    if self.BUS[b1][3] != 3:
                        # diagonal elements of J1
                        A[J1_diag_idx][J1_diag_idx] = J11        
                        DC[J1_diag_idx] = P[b1] - Pk
                    if self.BUS[b1][3] == 1:
                        # diagonal elements of J2
                        A[J2_row_diag_idx][J2_col_diag_idx] = 2 * Vm[b1] * Ym[b1] * math.cos(theta[b1]) + J22    
                        # diagonal elements of J3
                        A[J3_row_diag_idx][J3_col_diag_idx] = J33
                        # diagonal elements of J4         
                        A[J4_diag_idx][J4_diag_idx] = -2 * Vm[b1] * Ym[b1] * math.sin(theta[b1]) - J44   
                        DC[J4_diag_idx] = Q[b1] - Qk
                #matrix A left division Matrix DC
                DX = np.linalg.solve(A, DC.T)
                for bi in self.setBusHnd:
                    del_update_DXidx = int(bi - slackCounted[bi-1]-1)
                    Vm_update_DXidx = int(self.nBus + bi - PVcounted[bi-1] - slackCounted[bi-1] - self.nSlack-1)
                    if self.BUS[bi][3] != 3:
                        delta[bi] +=  DX[del_update_DXidx]
                    if self.BUS[bi][3] == 1:
                        Vm[bi] += DX[Vm_update_DXidx]
                maxerror = max(abs(DC))
                if maxerror < self.epsilon:
                    break
                if ii==self.iterMax:
                    return {'FLAG':'DIVERGENCE'}
            # finish NR
            # store all voltage magnitude value in all profile
            va.extend(Vm.values())
            vbus = {bi:(Vm[bi]* np.cos(delta[bi])+Vm[bi]*1j*np.sin(delta[bi])) for bi in self.setBusHnd }
            sbus = {bi:(P[bi] + Q[bi]*1j) for bi in self.setBusHnd}
            for bi in self.busSlack:
                sbus[bi] += self.loadProfile[pi][bi]
            """
            for bi in self.busPV:
                # update Q load in PV bus
                sbus[bi] += self.loadProfile[pi][bi].imag * 1j 
            """
            slt = 0
            Il = dict()
            for li,bi in self.lineC.items():
                line = frozenset({bi[0],bi[1]})
                Ib1 = (vbus[bi[0]]-vbus[bi[1]])*(-Ybus[line])
                Ib2 = (vbus[bi[1]]-vbus[bi[0]])*(-Ybus[line])
                #
                if abs(Ib1) >= abs(Ib2):
                    Il[li] = abs(Ib1)
                else: 
                    Il[li] = abs(Ib2)   
                rate = ((Il[li])/self.LINE[li][3])*RATEC
                ra.append(rate)
                
                Snk = vbus[bi[0]]*np.conj(Ib1)
                Skn = vbus[bi[1]]*np.conj(Ib2)
                slt += Snk +Skn 
            if fo:     
                res,cosP,cosN,rB,rL,rG =  self.__update1profilevalue__(pi,fo,res,va1,dia1,sa1,vbus,Il
                                                                    ,sbus,cosP,cosN,rB,rL,rG,slt)
            else:
                res,cosP,cosN = self.__update1profilevalue__(pi,fo,res,va1,dia1,sa1,vbus,Il
                                                                    ,sbus,cosP,cosN,slt)        
        #
        res['Umax[pu]'] = max(va)/self.Ubase
        res['Umin[pu]'] = min(va)/self.Ubase
        res['RateMax[%]'] = max(ra)
        res['cosP'] = min(cosP)
        res['cosN'] = max(cosN) 
        #
        if fo:
            rB.append(['','Umax[pu]',toString(res['Umax[pu]']),'Umin[pu]',toString(res['Umin[pu]']) ])
            add2CSV(fo,rB,',')
            #
            rL.append(['','RateMax[%]',toString(res['RateMax[%]'],2)])
            add2CSV(fo,rL,',')
            #
            rG.append(['','cosPmin',toString(res['cosP'],3),'cosNMax',toString(res['cosN'],3)])
            add2CSV(fo,rG,',')
        #
        return res     

    def __run1configNR__(self,lineOff,shuntOff,fo=''):
        # ready to run Newton raphson
        Ybus = self.__calculateYbus__(lineOff,shuntOff)
        
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
        #
        if fo:
            add2CSV(fo,[[],[time.ctime()],['PF 1Profile','lineOff',str(list(lineOff)),'shuntOff',str(list(shuntOff))]],',')
            #
            rB = [[],['BUS/Profile']]
            rB[1].extend([bi for bi in self.lstBusHnd])
            #
            rL = [[],['LINE/Profile']]
            rL[1].extend([bi for bi in self.lstLineHnd])
            #
            rG = [[],['GEN/Profile']]
            for bi in self.busSlack:
                rG[1].append(str(bi)+'_P')
                rG[1].append(str(bi)+'_Q')
                rG[1].append(str(bi)+'_cosPhi')
        #
        # return Y bus magnitude
        Ym = np.abs(Ybus)
        # return phase angle of Ybus        
        theta = np.angle(Ybus,deg=False)
        countSlack = 0
        countPV = 0
        slackCounted = []
        PVcounted = []
        for bi in self.BUS.keys():
            if self.BUS[bi][3] == 3:
                countSlack+=1 
            elif self.BUS[bi][3] == 2:
                countPV += 1 
            slackCounted.append(countSlack)
            PVcounted.append(countPV)
        no_jacobi_equation = 2 * self.nBus - countPV - 2 * self.nSlack
        #
        DC = np.zeros(no_jacobi_equation)
        #
        va,ra,cosP,cosN = [],[],[1],[-1]
        for pi in self.profileID:
            # initialize voltage magnitude of all buses
            Vm = [float(self.Ubase) for _ in self.BUS.keys()]
            P = [(-self.loadProfile[pi][bi]).real for bi in self.BUS.keys()]    #+self.Pgen
            Q = [(-self.loadProfile[pi][bi]).imag for bi in self.BUS.keys()]    #+self.Qgen
            # update Vm of slack buses
            for bs in self.setSlack:
                Vm[bs-1] = self.genProfile[pi][bs]
            # initialize voltage angle 
            delta = [0 for _ in self.BUS.keys()]
            sa1,dia1,va1 = dict(),dict(),dict()# for 1 profile
            for ii in range(self.iterMax+1):
                A = np.zeros((no_jacobi_equation,no_jacobi_equation))
                for b1 in self.BUS.keys():
                    J1_row_offdiag_idx = J2_row_offdiag_idx = int((b1 - slackCounted[b1-1])-1)
                    J1_diag_idx = J2_row_diag_idx = J3_col_diag_idx = J1_row_offdiag_idx 

                    J3_row_offdiag_idx = J4_row_offdiag_idx = int((self.nBus+b1-slackCounted[b1-1]-PVcounted[b1-1]-self.nSlack)-1)
                    J4_diag_idx = J2_col_diag_idx = J3_row_diag_idx = J3_row_offdiag_idx 

                    J11 = 0
                    J22 = 0
                    J33 = 0
                    J44 = 0
                    #
                    for li in self.busC[b1]:
                        for b2 in self.lineC[li]:
                            if b1 == b2:
                                pass
                            else:
                                # diagonal elements of J1
                                J11 += Vm[b1-1] * Vm[b2-1] * Ym[b1-1][b2-1] * math.sin((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                # diagonal elements of J3          
                                J33 += Vm[b1-1] * Vm[b2-1] * Ym[b1-1][b2-1] * math.cos((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                if self.BUS[b1][3] != 3:
                                    #slackbus doesn't exist in J2 and J4
                                    #diagonal elements of J2
                                    J22 += Vm[b2-1] * Ym[b1-1][b2-1] * math.cos((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                    #diagonal elements of J4               
                                    J44 += Vm[b2-1] * Ym[b1-1][b2-1] * math.sin((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)

                                if self.BUS[b1][3] != 3 and self.BUS[b2][3] != 3:
                                    J1_col_offdiag_idx = J3_col_offdiag_idx = int(b2 - slackCounted[b2-1] - 1)
                                    J2_col_offdiag_idx = J4_col_offdiag_idx = int(self.nBus + b2 - PVcounted[b2-1] - slackCounted[b2-1] - self.nSlack - 1)
                                    # off diagonal elements of J1
                                    A[J1_row_offdiag_idx][J1_col_offdiag_idx] = -Vm[b1-1] * Vm[b2-1] * Ym[b1-1][b2-1] * math.sin((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                
                                    if self.BUS[b2][3] == 1:
                                        # off diagonal elements of J2
                                        A[J2_row_offdiag_idx][J2_col_offdiag_idx] = Vm[b1-1] * Ym[b1-1][b2-1] * math.cos((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                
                                    if self.BUS[b1][3] == 1:
                                        # off diagonal elements of J3
                                        A[J3_row_offdiag_idx][J3_col_offdiag_idx] = -Vm[b1-1] * Vm[b2-1] * Ym[b1-1][b2-1] * math.cos((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                                
                                    if self.BUS[b1][3] == 1 and self.BUS[b2][3] == 1:
                                        # off diagonal elements of J4
                                        A[J4_row_offdiag_idx][J4_col_offdiag_idx] = -Vm[b1-1] * Ym[b1-1][b2-1] * math.sin((theta[b1-1][b2-1] - delta[b1-1] + delta[b2-1]).real)
                    #   
                    Pk = Vm[b1-1]**2 * Ym[b1-1][b1-1] * math.cos((theta[b1-1][b1-1])) + J33
                    Qk = -Vm[b1-1]**2 * Ym[b1-1][b1-1] * math.sin((theta[b1-1][b1-1])) - J11
                    if self.BUS[b1][3] == 3:
                        # swing bus
                        P[b1-1] = Pk
                        Q[b1-1] = Qk
                    if self.BUS[b1][3] == 2:
                        Q[b1-1] = Qk 
                        """
                        if Qmax[b1-1] != 0:
                            Qgc = Q[b1-1] + Qd[n-1] - Qsh[n-1]
                            if ii <= 7:                   #Between the 2th & 6th iterations
                                if ii > 2:                #the Mvar of generator buses are
                                    if Qgc < Qmin[n-1]:   #tested. If not within limits Vm(n)
                                        Vm[n-1] += 0.01   #is changed in steps of 0.01 pu to
                                    elif Qgc > Qmax[n-1]: #bring the generator Mvar within
                                        Vm[n-1] -= 0.01   #the specified limits.            
                        """
                    if self.BUS[b1][3] != 3:
                        #diagonal elements of J1
                        A[J1_diag_idx][J1_diag_idx] = J11        
                        DC[J1_diag_idx] = P[b1-1] - Pk
                    if self.BUS[b1][3] == 1:
                        #diagonal elements of J2
                        A[J2_row_diag_idx][J2_col_diag_idx] = 2 * Vm[b1-1] * Ym[b1-1][b1-1] * math.cos(theta[b1-1][b1-1]) + J22    
                        #diagonal elements of J3
                        A[J3_row_diag_idx][J3_col_diag_idx] = J33
                        #diagonal elements of J4         
                        A[J4_diag_idx][J4_diag_idx] = -2 * Vm[b1-1] * Ym[b1-1][b1-1] * math.sin(theta[b1-1][b1-1]) - J44   
                        DC[J4_diag_idx] = Q[b1-1] - Qk
                #matrix A left division Matrix DC
                DX = np.linalg.solve(A, DC.T)
                for bi in self.BUS.keys():
                    del_update_DXidx = int(bi - slackCounted[bi-1]-1)
                    Vm_update_DXidx = int(self.nBus + bi - PVcounted[bi-1] - slackCounted[bi-1] - self.nSlack-1)
                    if self.BUS[bi][3] != 3:
                        delta[bi-1] +=  DX[del_update_DXidx]
                    if self.BUS[bi][3] == 1:
                        Vm[bi-1] += DX[Vm_update_DXidx]
                maxerror = max(abs(DC))
                if maxerror < self.epsilon:
                    break
                if ii==self.iterMax:
                    return {'FLAG':'DIVERGENCE'}
            #finish NR
            # store all voltage magnitude value in all profile
            va.extend(Vm)
            V = [Vm[bi-1]* np.cos(delta[bi-1])+Vm[bi-1]*1j*np.sin(delta[bi-1]) for bi in self.busC ]
            Il = dict()
            SLT = 0
            vbus,sbus = dict(),dict()
            for i in range(len(Vm)):
                vbus[i+1] = V[i]
                sbus[i+1] = P[i] + Q[i]*1j
            for bi in self.busSlack:
                sbus[bi] += self.loadProfile[pi][bi]
            for li,bi in self.lineC.items():
                Ib1 = (vbus[bi[0]]-vbus[bi[1]])*(-Ybus[bi[0]-1][bi[1]-1])
                Ib2 = (vbus[bi[1]]-vbus[bi[0]])*(-Ybus[bi[0]-1][bi[1]-1])
                #
                if abs(Ib1) >= abs(Ib2):
                    Il[li] = abs(Ib1)
                else: 
                    Il[li] = abs(Ib2)   

                 
                rate = ((Il[li])/self.LINE[li][3])*RATEC
                ra.append(rate)
                
                Snk = V[bi[0]-1]*np.conj(Ib1)
                Skn = V[bi[1]-1]*np.conj(Ib2)
                SLT += Snk +Skn  
                
            if fo:
                va1.update(vbus)
                dia1.update(Il)
                sa1.update(sbus)
            for bs in self.busSlack:
                if sbus[bs].imag:
                    cosP.append(sbus[bs].real/abs(sbus[bs]))
                else:
                    cosN.append(-sbus[bs].real/abs(sbus[bs]))
            if fo:
                rb1 = [pi]
                rl1 = [pi]
                rg1 = [pi]
                for bi1 in self.lstBusHnd:
                    rb1.append(toString(abs(va1[bi1])/self.Ubase))
                #
                for bri in self.lstLineHnd:
                    try:
                        r1 = abs(Il[bri])/self.LINE[bri][3]*RATEC
                        rl1.append( toString(r1,2) )
                    except:
                        rl1.append('0')
                #
                for bs1 in self.busSlack:
                    rg1.append(toString(sa1[bs1].real))
                    rg1.append(toString(sa1[bs1].imag))
                    if sa1[bs1].imag>=0:
                        rg1.append(toString(sa1[bs1].real/abs(sa1[bs1]),3))
                    else:
                        rg1.append(toString(-sa1[bs1].real/abs(sa1[bs1]),3))
                #
                rB.append(rb1)
                rL.append(rl1)
                rG.append(rg1)
            res['DeltaA'] += SLT.real

        res['Umax[pu]'] = max(va)/self.Ubase
        res['Umin[pu]'] = min(va)/self.Ubase
        res['RateMax[%]'] = max(ra)
        res['cosP'] = min(cosP)
        res['cosN'] = max(cosN)
        
        #
        if fo:
            rB.append(['','Umax[pu]',toString(res['Umax[pu]']),'Umin[pu]',toString(res['Umin[pu]']) ])
            add2CSV(fo,rB,',')
            #
            rL.append(['','RateMax[%]',toString(res['RateMax[%]'],2)])
            add2CSV(fo,rL,',')
            #
            rG.append(['','cosPmin',toString(res['cosP'],3),'cosNMax',toString(res['cosN'],3)])
            add2CSV(fo,rG,',')
        #
        return res        
#
class GaussSeidel(POWERFLOW):
    def __init__(self, fi):
        super().__init__(fi)
    def __run1configGS__(self,lineOff,shuntOff,fo=''):
        # ready to run Gauss Seidel
        Ybus = self.__calculate_sparse_Ybus__(lineOff,shuntOff)
        #
        res = {'FLAG':'CONVERGENCE','RateMax[%]':0, 'Umax[pu]':0,'Umin[pu]':100,'DeltaA':0,'cosP':0,'cosN':0}
         #
        if fo:
            add2CSV(fo,[[],[time.ctime()],['PF 1Profile','lineOff',str(list(lineOff)),'shuntOff',str(list(shuntOff))]],',')
            #
            rB = [[],['BUS/Profile']]
            rB[1].extend([bi for bi in self.lstBusHnd])
            #
            rL = [[],['LINE/Profile']]
            rL[1].extend([bi for bi in self.lstLineHnd])
            #
            rG = [[],['GEN/Profile']]
            for bi in self.busSlack:
                 rG[1].append(str(bi)+'_P')
                 rG[1].append(str(bi)+'_Q')
                 rG[1].append(str(bi)+'_cosPhi')
        va,ra,cosP,cosN = [],[],[1],[-1]
        accel = self.setting['accel']
        for pi in self.profileID:
            P = {bi:(-self.loadProfile[pi][bi]).real for bi in self.setBusHnd}    #+self.Pgen
            Q = {bi:(-self.loadProfile[pi][bi]).imag for bi in self.setBusHnd}    #+self.Qgen
            DP = {bi:0 for bi in self.setBusHnd}
            DQ = {bi:0 for bi in self.setBusHnd}
            sbus = {bi:(P[bi] + Q[bi]*1j) for bi in self.setBusHnd}
            vbus = {bi:complex(self.Ubase,0) for bi in self.setBusHnd}
            # initialize voltage correction
            Vc = {bi:complex(0,0) for bi in self.setBusHnd}
            # update Vm of slack buses
            for bs in self.setSlack:
                vbus[bs] = complex(self.genProfile[pi][bs],0) 
            Vm = {bi:abs(vbus[bi])for bi in self.setBusHnd}
            sa1,dia1,va1 = dict(),dict(),dict() # for 1 profile
            for ii in range(self.iterMax+1):
                for b1 in self.setBusHnd:
                    YV = 0 + 1j * 0
                    for li in self.busC[b1]:
                        for b2 in self.lineC[li]:
                            if b1 == b2:
                                pass
                            else:
                                line = frozenset({b1,b2})
                                YV += Ybus[line] * vbus[b2]
                    Sc = np.conj(vbus[b1]) * (Ybus[b1] * vbus[b1] + YV)
                    Sc = np.conj(Sc)
                    DP[b1] = P[b1] - np.real(Sc)
                    DQ[b1] = Q[b1] - np.imag(Sc)
                    if self.BUS[b1][3] == 3:
                        sbus[b1] = Sc
                        P[b1] = np.real(Sc)
                        Q[b1] = np.imag(Sc)
                        DP[b1] = 0
                        DQ[b1] = 0
                        Vc[b1] = vbus[b1]
                    elif self.BUS[b1][3] == 2:
                        Q[b1] = np.imag(Sc)
                        sbus[b1] = P[b1] + 1j * Q[b1]
                        """
                        if Qmax[n] != 0:
                            Qgc = Q[n] * basemva + Qd[n] - Qsh[n]
                            if abs(DQ[n]) <= 0.005 and iter >= 10:
                                if DV[n] <= 0.045:
                                    if Qgc < Qmin[n]:
                                        Vm[n] += 0.005
                                        DV[n] += 0.005
                                    elif Qgc > Qmax[n]:
                                        Vm[n] -= 0.005
                                        DV[n] += 0.005
                                else:
                                    pass
                            else:
                                pass
                        else:
                            pass
                        """
                    if self.BUS[b1][3] != 3:
                        Vc[b1] = (np.conj(sbus[b1]) / np.conj(vbus[b1]) - YV) / Ybus[b1]
                    else:
                        pass
                    if self.BUS[b1][3] == 1:
                        vbus[b1] = vbus[b1] + accel * (Vc[b1] - vbus[b1])
                    elif self.BUS[b1][3] == 2:
                        VcI = np.imag(Vc[b1])
                        VcR = np.sqrt(Vm[b1] ** 2 - VcI ** 2)
                        Vc[b1] = VcR + 1j * VcI
                        vbus[b1] = vbus[b1] + accel * (Vc[b1] - vbus[b1])
                maxerror = abs(DP[b1])
                for bi in self.setBusHnd:
                    if abs(DP[bi]) > maxerror:
                        maxerror = abs(DP[bi])
                    if abs(DQ[bi]) > maxerror:
                        maxerror = abs(DQ[bi])
                if maxerror < self.epsilon:
                    break
                if ii==self.iterMax:
                    return {'FLAG':'DIVERGENCE'}
            # finish GS
            # store all voltage magnitude value in all profile
            Vm = {bi:abs(vbus[bi]) for bi in self.setBusHnd}
            va.extend(Vm.values())
            for bi in self.busSlack:
                sbus[bi] += self.loadProfile[pi][bi]
            """
            for bi in self.busPV:
                # update Q load in PV bus
                sbus[bi] += self.loadProfile[pi][bi].imag * 1j 
            """
            slt = 0
            Il = dict()
            for li,bi in self.lineC.items():
                line = frozenset({bi[0],bi[1]})
                Ib1 = (vbus[bi[0]]-vbus[bi[1]])*(-Ybus[line])
                Ib2 = (vbus[bi[1]]-vbus[bi[0]])*(-Ybus[line])
                #
                if abs(Ib1) >= abs(Ib2):
                    Il[li] = abs(Ib1)
                else: 
                    Il[li] = abs(Ib2)   

                rate = ((Il[li])/self.LINE[li][3])*RATEC
                ra.append(rate)
                
                Snk = vbus[bi[0]]*np.conj(Ib1)
                Skn = vbus[bi[1]]*np.conj(Ib2)
                slt += Snk +Skn
            
            res,cosP,cosN,rB,rL,rG =  self.__update1profilevalue__(pi,fo,res,va1,dia1,sa1,vbus,Il
                                                                    ,sbus,cosP,cosN,rB,rL,rG,slt)
        #
        res['Umax[pu]'] = max(va)/self.Ubase
        res['Umin[pu]'] = min(va)/self.Ubase
        res['RateMax[%]'] = max(ra)
        res['cosP'] = min(cosP)
        res['cosN'] = max(cosN)
        
        #
        if fo:
            rB.append(['','Umax[pu]',toString(res['Umax[pu]']),'Umin[pu]',toString(res['Umin[pu]']) ])
            add2CSV(fo,rB,',')
            #
            rL.append(['','RateMax[%]',toString(res['RateMax[%]'],2)])
            add2CSV(fo,rL,',')
            #
            rG.append(['','cosPmin',toString(res['cosP'],3),'cosNMax',toString(res['cosN'],3)])
            add2CSV(fo,rG,',')
        #
        return res      
    
# #
def test_psm():
    # 1 source

    #ARGVS.fi = 'Inputs12.xlsx'
##    varFlag = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 12, 13, 14, 15, 16,0,1]
    #lineOff = [12,13,14,15,16]
    #shuntOff = []

##    # 2 source
    ARGVS.fi = 'Inputs12_2.xlsx'
    #varFlag = [0, 1, 1, 0, 0, 1, 0, 0, 0, 0, 1, 1, 0, 1, 0, 0, 1]
    #lineOff = [6,7,10,11,13,16]
    #lineOff = [3,12,13,14,16]
    lineOff = [3,12,13,14,15,16]
#    lineOff = [6,12,14]
    shuntOff = [0]
    #lineOff = []
    #shuntOff = []
    # 190 bus
    #ARGVS.fi = 'Inputs190.xlsx'
    #lineOff = [66, 103, 110, 169, 191]
    #shuntOff = [47, 66, 80, 130]
    #ARGVS.fi = 'Inputs33bus.xlsx'
    #
    p1 = POWERFLOW(ARGVS.fi)
    t01 = time.time()
    v1 = p1.run1Config_WithObjective(lineOff=lineOff,shuntOff=shuntOff,fo=ARGVS.fo)
##    print(v1)
    #v1 = p1.run1Config_WithObjective(lineOff=lineOff,shuntOff=shuntOff,fo=ARGVS.fo)
    print('time %.5f'%(time.time()-t01))
    print(v1)
#



if __name__ == '__main__':
    
    ARGVS.fo = 'res\\res1Config.csv'
    test_psm()
    """
    ARGVS.fi = 'Inputs12_2.xlsx'
    p1 = POWERFLOW(ARGVS.fi)
    lineOff = [6,7,10,11,13,16]
    fl1 = p1.__checkLoopIsland__(lineOff=set(lineOff))
    print(fl1)
    """